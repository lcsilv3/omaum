from openpyxl.styles import Font, PatternFill


class ExcelGenerator:
    """Gerador de relatórios Excel com fidelidade visual."""

    def __init__(self, template_path):
        self.template_path = template_path

    def gerar_consolidado(self, dados):
        """Gera relatório consolidado em Excel."""
        # This is a placeholder for loading a template. In a real scenario, the template
        # would be a pre-formatted Excel file.
        # For now, we create a new workbook.
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Consolidado"

        # Apply styles and fill data
        self._aplicar_estilos_cabecalho(ws)
        self._preencher_dados_consolidado(ws, dados)
        self._aplicar_formatacao_condicional(ws, dados)

        return wb

    def _aplicar_estilos_cabecalho(self, ws):
        """Aplica estilos ao cabeçalho."""
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        # Basic headers
        headers = [
            "Aluno",
            "Mês/Ano",
            "Presentes",
            "Faltas",
            "Justificadas",
            "Voluntário Extra",
            "Voluntário Simples",
            "Total Presença (%)",
        ]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill

    def _preencher_dados_consolidado(self, ws, dados):
        """Preenche a planilha com os dados processados."""
        row = 5
        for aluno_dados in dados["alunos"]:
            # This is a simplified representation. The proposal implies a more complex pivot table structure.
            # Here we will list totals per student.
            totais = aluno_dados["totais"]
            total_atividades = (
                totais.get("P", 0) + totais.get("F", 0) + totais.get("J", 0)
            )
            percentual_presenca = (
                (totais.get("P", 0) / total_atividades * 100)
                if total_atividades > 0
                else 0
            )

            ws.cell(row=row, column=1, value=aluno_dados["aluno"].nome)
            # Column 2 (Mês/Ano) is skipped in this simplified total view
            ws.cell(row=row, column=3, value=totais.get("P", 0))
            ws.cell(row=row, column=4, value=totais.get("F", 0))
            ws.cell(row=row, column=5, value=totais.get("J", 0))
            ws.cell(row=row, column=6, value=totais.get("V1", 0))
            ws.cell(row=row, column=7, value=totais.get("V2", 0))
            ws.cell(row=row, column=8, value=f"{percentual_presenca:.2f}%")
            row += 1

    def _calcular_percentual_presenca(self, aluno_dados):
        totais = aluno_dados["totais"]
        total_convocacoes = totais.get("P", 0) + totais.get("F", 0) + totais.get("J", 0)
        if total_convocacoes == 0:
            return 0
        return (totais.get("P", 0) / total_convocacoes) * 100

    def _aplicar_formatacao_condicional(self, ws, dados):
        """Aplica formatação condicional para baixa presença."""
        red_fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )

        for row_idx, aluno_dados in enumerate(dados["alunos"], start=5):
            percentual = self._calcular_percentual_presenca(aluno_dados)

            if percentual < 75:  # Critério de baixa presença
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).fill = red_fill
