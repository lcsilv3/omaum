#!/usr/bin/env python
"""Sincroniza os cursos cadastrados a partir de planilha CSV ou XLSX."""

from __future__ import annotations

import argparse
import csv
import os
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "omaum.settings.development")

import django  # noqa: E402  # pylint: disable=wrong-import-position


django.setup()

from cursos import services  # noqa: E402  # pylint: disable=wrong-import-position

DOCS_DIR = PROJECT_ROOT / "docs"
PLANILHAS_CANDIDATAS = (
    DOCS_DIR / "Planilha de Cursos.csv",
    DOCS_DIR / "Planilha de Cursos.xlsx",
)


def localizar_primeiro_existente(caminhos: Iterable[Path]) -> Optional[Path]:
    """Retorna o primeiro caminho existente dentro da sequência fornecida."""

    for caminho in caminhos:
        if caminho.exists():
            return caminho
    return None


def _normalizar_coluna(chave: Any) -> str:
    texto = services.normalizar_texto(chave)
    base = (
        unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    )
    return base.replace(" ", "").replace("_", "").lower()


def _linha_vazia(row: Dict[str, Any]) -> bool:
    return not any(services.normalizar_texto(valor) for valor in row.values())


def _mapear_registro(row: Dict[str, Any], linha_planilha: int) -> Dict[str, Any]:
    colunas_normalizadas = {
        _normalizar_coluna(coluna): valor for coluna, valor in row.items() if coluna
    }

    return {
        "linha": linha_planilha,
        "id": colunas_normalizadas.get("id") or colunas_normalizadas.get("codigo"),
        "nome": colunas_normalizadas.get("nome")
        or colunas_normalizadas.get("curso")
        or colunas_normalizadas.get("titulo"),
        "descricao": colunas_normalizadas.get("descricao")
        or colunas_normalizadas.get("descricaodocurso"),
        "ativo": colunas_normalizadas.get("ativo")
        or colunas_normalizadas.get("status"),
    }


def _carregar_csv(caminho: Path) -> List[Dict[str, Any]]:
    with caminho.open("r", encoding="utf-8-sig", newline="") as ponteiro:
        amostra = ponteiro.read(2048)
        ponteiro.seek(0)
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","
        leitor = csv.DictReader(ponteiro, delimiter=delimitador)
        registros: List[Dict[str, Any]] = []
        for indice, row in enumerate(leitor, start=2):
            if _linha_vazia(row):
                continue
            registros.append(_mapear_registro(row, indice))
        return registros


def _carregar_xlsx(caminho: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependencia opcional
        raise RuntimeError(
            "openpyxl e necessario para ler arquivos .xlsx; instale a dependencia"
        ) from exc

    workbook = load_workbook(filename=caminho, read_only=True, data_only=True)
    sheet = workbook.active
    headers: List[str] = []
    registros: List[Dict[str, Any]] = []

    for indice, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        valores = [services.normalizar_texto(celula) for celula in row]
        if indice == 1:
            headers = valores
            continue
        row_dict: Dict[str, Any] = {
            headers[posicao] if posicao < len(headers) else f"coluna_{posicao}": valor
            for posicao, valor in enumerate(valores)
        }
        if _linha_vazia(row_dict):
            continue
        registros.append(_mapear_registro(row_dict, indice))

    workbook.close()
    return registros


def carregar_planilha(caminho: Path) -> List[Dict[str, Any]]:
    """Carrega a planilha de cursos para uma lista de dicionários normalizados."""

    sufixo = caminho.suffix.lower()
    if sufixo == ".csv":
        return _carregar_csv(caminho)
    if sufixo in {".xlsx", ".xlsm"}:
        return _carregar_xlsx(caminho)
    raise ValueError(f"Formato de planilha não suportado: {sufixo}")


def imprimir_resumo(caminho: Path, resumo: Dict[str, Any]) -> None:
    """Imprime contagens e avisos obtidos na sincronização."""

    print("\nSincronizacao de Cursos")
    print(f"Fonte: {caminho}")
    print(
        "Resumo: processados={processados}, criados={criados}, "
        "atualizados={atualizados}, reativados={reativados}, desativados={desativados}".format(
            **resumo
        )
    )
    avisos = resumo.get("avisos", [])
    if avisos:
        print("Avisos (até 10 exibidos):")
        for aviso in avisos[:10]:
            print(f" - {aviso}")
        if len(avisos) > 10:
            print(" - ... (demais avisos omitidos)")


def executar_sincronizacao(arquivo: Optional[Path], manter_existentes: bool) -> None:
    """Executa a carga dos dados e aplica a sincronização."""

    caminho = arquivo or localizar_primeiro_existente(PLANILHAS_CANDIDATAS)
    if not caminho:
        print(
            "Nenhum arquivo de cursos encontrado. Coloque a planilha em docs/ e tente novamente."
        )
        return

    registros = carregar_planilha(caminho)
    if not registros:
        print("Planilha sem registros válidos. Nada foi alterado.")
        return

    resumo = services.sincronizar_cursos(
        registros,
        desativar_nao_listados=not manter_existentes,
    )
    imprimir_resumo(caminho, resumo)


def obter_argumentos() -> argparse.Namespace:
    """Lê argumentos de linha de comando."""

    parser = argparse.ArgumentParser(
        description="Sincroniza os cursos cadastrados a partir da planilha oficial.",
    )
    parser.add_argument(
        "--arquivo",
        type=Path,
        help="Caminho alternativo para a planilha CSV ou XLSX de cursos.",
    )
    parser.add_argument(
        "--manter-existentes",
        action="store_true",
        help="Não desativa cursos que não estejam presentes na planilha.",
    )
    return parser.parse_args()


def main() -> None:
    """Função principal do script CLI."""

    argumentos = obter_argumentos()
    executar_sincronizacao(argumentos.arquivo, argumentos.manter_existentes)


if __name__ == "__main__":
    main()
