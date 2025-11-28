#!/usr/bin/env python
"""Sincroniza tipos e códigos iniciáticos a partir das planilhas oficiais."""

from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple

# Configuração mínima de ambiente Django ---------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "omaum.settings.development")
import django  # noqa: E402  # pylint: disable=wrong-import-position

django.setup()

from alunos.utils import get_codigo_model, get_tipo_codigo_model  # noqa: E402

Codigo: Any = get_codigo_model()
TipoCodigo: Any = get_tipo_codigo_model()
if not (Codigo and TipoCodigo):  # pragma: no cover - verificação defensiva
    raise RuntimeError("Modelos iniciáticos (Codigo/TipoCodigo) indisponíveis.")

DOCS_DIR = PROJECT_ROOT / "docs"
PLANILHA_TIPOS_CANDIDATOS = (
    DOCS_DIR / "Planilha Tipos de  Códigos.csv",
    DOCS_DIR / "Planilha Tipos de Códigos.csv",
    DOCS_DIR / "Planilha Tipos de Códigos.xlsx",
)
PLANILHA_CODIGOS_CANDIDATOS = (
    DOCS_DIR / "Planilha de Códigos.csv",
    DOCS_DIR / "Planilha de Códigos.xlsx",
)


# Utilidades --------------------------------------------------------------------


def localizar_primeiro_existente(caminhos: Iterable[Path]) -> Optional[Path]:
    """Retorna o primeiro caminho existente dentro da sequência fornecida."""

    for caminho in caminhos:
        if caminho.exists():
            return caminho
    return None


def normalizar_texto(valor: Any) -> str:
    """Converte o valor para string e aplica strip defensivo."""

    if valor is None:
        return ""
    return str(valor).strip()


def converter_para_int(valor: Any) -> Optional[int]:
    """Converte um valor textual em inteiro tolerando sufixos ".0"."""

    texto = normalizar_texto(valor)
    if not texto:
        return None
    if texto.endswith(".0"):
        texto = texto[:-2]
    try:
        return int(texto)
    except ValueError:
        try:
            return int(float(texto))
        except (TypeError, ValueError):
            return None


def carregar_planilha(caminho: Path) -> List[List[str]]:
    """Carrega dados CSV/XLSX para lista de linhas com strings normalizadas."""

    if caminho.suffix.lower() == ".csv":
        with caminho.open("r", encoding="utf-8-sig", newline="") as ponteiro:
            leitor = csv.reader(ponteiro, delimiter=";")
            return [[normalizar_texto(celula) for celula in linha] for linha in leitor]

    if caminho.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover - dependência opcional
            raise RuntimeError(
                "openpyxl é necessário para ler arquivos .xlsx; instale a dependência"
            ) from exc

        workbook = load_workbook(filename=caminho, read_only=True, data_only=True)
        planilha = workbook.active
        linhas: List[List[str]] = []
        for linha in planilha.iter_rows(values_only=True):
            linhas.append([normalizar_texto(celula) for celula in linha])
        return linhas

    raise ValueError(f"Formato de planilha não suportado: {caminho.suffix}")


def remover_linhas_vazias(linhas: Iterable[List[str]]) -> Iterator[List[str]]:
    """Remove linhas sem conteúdo para evitar processamento desnecessário."""

    for linha in linhas:
        if any(celula for celula in linha):
            yield linha


# Sincronização de Tipos ---------------------------------------------------------


def sincronizar_tipos(
    arquivo: Optional[Path] = None,
) -> Tuple[Dict[int, Any], Dict[str, Any]]:
    """Sincroniza a planilha de tipos e devolve mapa id->objeto com resumo."""

    caminho = arquivo or localizar_primeiro_existente(PLANILHA_TIPOS_CANDIDATOS)
    resumo = {
        "arquivo": str(caminho) if caminho else None,
        "criados": 0,
        "atualizados": 0,
        "reativados": 0,
        "desativados": 0,
        "avisos": [],
    }

    if not caminho:
        resumo["avisos"].append("Arquivo de tipos não encontrado.")
        return {}, resumo

    linhas = carregar_planilha(caminho)
    linhas_util = list(remover_linhas_vazias(linhas))
    if len(linhas_util) <= 1:
        resumo["avisos"].append("Planilha de tipos vazia ou sem registros válidos.")
        return {}, resumo

    tipos_por_id: Dict[int, Any] = {}
    ids_importados: List[int] = []

    for indice, linha in enumerate(linhas_util[1:], start=2):
        tipo_id = converter_para_int(linha[0] if len(linha) > 0 else None)
        nome = normalizar_texto(linha[1] if len(linha) > 1 else None)
        descricao = normalizar_texto(linha[2] if len(linha) > 2 else None)

        if tipo_id is None:
            resumo["avisos"].append(f"Linha {indice}: identificador do tipo ausente.")
            continue
        if not nome:
            resumo["avisos"].append(
                f"Linha {indice}: nome não informado para o tipo {tipo_id}."
            )
            continue

        existente = TipoCodigo.objects.filter(id=tipo_id).first()
        estava_inativo = bool(existente and not existente.ativo)

        objeto, criado = TipoCodigo.objects.update_or_create(
            id=tipo_id,
            defaults={"nome": nome, "descricao": descricao or None, "ativo": True},
        )
        if criado:
            resumo["criados"] += 1
        else:
            resumo["atualizados"] += 1
            if estava_inativo:
                resumo["reativados"] += 1

        objeto.refresh_from_db()
        tipos_por_id[objeto.id] = objeto
        ids_importados.append(objeto.id)

    if ids_importados:
        desativados = (
            TipoCodigo.objects.filter(ativo=True)
            .exclude(id__in=ids_importados)
            .update(ativo=False)
        )
        resumo["desativados"] = desativados

    return tipos_por_id, resumo


# Sincronização de Códigos -------------------------------------------------------


def sincronizar_codigos(
    tipos_por_id: Dict[int, Any], arquivo: Optional[Path] = None
) -> Dict[str, Any]:
    """Sincroniza a planilha de códigos utilizando os tipos previamente importados."""

    caminho = arquivo or localizar_primeiro_existente(PLANILHA_CODIGOS_CANDIDATOS)
    resumo = {
        "arquivo": str(caminho) if caminho else None,
        "criados": 0,
        "atualizados": 0,
        "reativados": 0,
        "desativados": 0,
        "avisos": [],
        "divergencias_tipo": [],
    }

    if not caminho:
        resumo["avisos"].append("Arquivo de códigos não encontrado.")
        return resumo

    linhas = carregar_planilha(caminho)
    linhas_util = list(remover_linhas_vazias(linhas))
    if len(linhas_util) <= 1:
        resumo["avisos"].append("Planilha de códigos vazia ou sem registros válidos.")
        return resumo

    nomes_importados: List[str] = []

    for indice, linha in enumerate(linhas_util[1:], start=2):
        tipo_id = converter_para_int(linha[1] if len(linha) > 1 else None)
        tipo_nome_planilha = normalizar_texto(linha[2] if len(linha) > 2 else None)
        codigo_nome = normalizar_texto(linha[3] if len(linha) > 3 else None)
        descricao = normalizar_texto(linha[4] if len(linha) > 4 else None)

        if not codigo_nome:
            resumo["avisos"].append(
                f"Linha {indice}: código sem identificador; ignorado."
            )
            continue

        if tipo_id is None or tipo_id not in tipos_por_id:
            resumo["avisos"].append(
                f"Linha {indice}: tipo {tipo_id!r} inexistente; código {codigo_nome} ignorado."
            )
            continue

        tipo_obj = tipos_por_id[tipo_id]
        if tipo_nome_planilha and tipo_nome_planilha != tipo_obj.nome:
            resumo["divergencias_tipo"].append(
                f"Linha {indice}: tipo informado '{tipo_nome_planilha}' difere do cadastrado '{tipo_obj.nome}'."
            )

        existente = Codigo.objects.filter(nome=codigo_nome).first()
        estava_inativo = bool(existente and not existente.ativo)

        objeto, criado = Codigo.objects.update_or_create(
            nome=codigo_nome,
            defaults={
                "tipo_codigo": tipo_obj,
                "descricao": descricao or None,
                "ativo": True,
            },
        )
        if criado:
            resumo["criados"] += 1
        else:
            resumo["atualizados"] += 1
            if estava_inativo:
                resumo["reativados"] += 1

        nomes_importados.append(objeto.nome)

    if nomes_importados:
        desativados = (
            Codigo.objects.filter(ativo=True)
            .exclude(nome__in=nomes_importados)
            .update(ativo=False)
        )
        resumo["desativados"] = desativados

    return resumo


# Saída formatada ----------------------------------------------------------------


def imprimir_resumo(titulo: str, resumo: Dict[str, Any]) -> None:
    """Imprime resumo amigável com contagens e avisos."""

    print(f"\n🗂️  {titulo}")
    if resumo.get("arquivo"):
        print(f"   • Fonte: {resumo['arquivo']}")
    for chave in ("criados", "atualizados", "reativados", "desativados"):
        if chave in resumo:
            print(f"   • {chave.capitalize()}: {resumo.get(chave, 0)}")

    if resumo.get("divergencias_tipo"):
        print("   • Divergências entre planilha e banco:")
        for aviso in resumo["divergencias_tipo"][:10]:
            print(f"     - {aviso}")
        if len(resumo["divergencias_tipo"]) > 10:
            print("     - ... (demais divergências omitidas)")

    if resumo.get("avisos"):
        print("   • Avisos:")
        for aviso in resumo["avisos"][:10]:
            print(f"     - {aviso}")
        if len(resumo["avisos"]) > 10:
            print("     - ... (demais avisos omitidos)")


# Execução -----------------------------------------------------------------------


def executar_pipeline(
    apenas_tipos: bool, arquivo_tipos: Optional[Path], arquivo_codigos: Optional[Path]
) -> None:
    """Executa o fluxo de sincronização conforme parâmetros informados."""

    tipos_por_id, resumo_tipos = sincronizar_tipos(arquivo_tipos)
    imprimir_resumo("Tipos de Código", resumo_tipos)

    if apenas_tipos:
        return

    if not tipos_por_id:
        print("\n⚠️  Nenhum tipo foi importado; sincronização de códigos cancelada.")
        return

    resumo_codigos = sincronizar_codigos(tipos_por_id, arquivo_codigos)
    imprimir_resumo("Códigos Iniciáticos", resumo_codigos)


def obter_argumentos() -> argparse.Namespace:
    """Interpreta argumentos de linha de comando."""

    parser = argparse.ArgumentParser(
        description="Sincroniza tipos e códigos iniciáticos a partir das planilhas oficiais.",
    )
    parser.add_argument(
        "--apenas-tipos",
        action="store_true",
        help="Processa apenas a planilha de tipos (não altera códigos).",
    )
    parser.add_argument(
        "--tipos-arquivo",
        type=Path,
        help="Caminho alternativo para a planilha de tipos.",
    )
    parser.add_argument(
        "--codigos-arquivo",
        type=Path,
        help="Caminho alternativo para a planilha de códigos.",
    )
    return parser.parse_args()


def main() -> None:
    """Função principal quando o script é executado via CLI."""

    argumentos = obter_argumentos()
    executar_pipeline(
        argumentos.apenas_tipos, argumentos.tipos_arquivo, argumentos.codigos_arquivo
    )


if __name__ == "__main__":
    main()
