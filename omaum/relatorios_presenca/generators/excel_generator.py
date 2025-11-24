"""
Gerador de relatórios Excel com fidelidade visual aos formatos existentes.

Este módulo implementa a geração de relatórios Excel que replicam
fielmente o layout e formatação das planilhas originais do projeto.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import io
import logging

logger = logging.getLogger(__name__)


class ExcelRelatorioGenerator:
    """
    Gerador de relatórios Excel com fidelidade visual.

    Implementa geração de relatórios que mantêm a aparência
    visual das planilhas Excel originais do projeto.
    """

    def __init__(self, template_path=None):
        """
        Inicializa o gerador.

        Args:
            template_path: Caminho para template Excel (opcional)
        """
        self.template_path = template_path
        self._configurar_estilos()

    def _configurar_estilos(self):
        """Configura estilos padrão para os relatórios."""
        # Estilo para cabeçalhos
        self.estilo_cabecalho = NamedStyle(name="cabecalho")
        self.estilo_cabecalho.font = Font(bold=True, color="FFFFFF", size=12)
        self.estilo_cabecalho.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        self.estilo_cabecalho.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.estilo_cabecalho.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Estilo para título
        self.estilo_titulo = NamedStyle(name="titulo")
        self.estilo_titulo.font = Font(bold=True, size=16, color="000000")
        self.estilo_titulo.alignment = Alignment(horizontal="center", vertical="center")

        # Estilo para dados
        self.estilo_dados = NamedStyle(name="dados")
        self.estilo_dados.font = Font(size=10)
        self.estilo_dados.alignment = Alignment(horizontal="center", vertical="center")
        self.estilo_dados.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Estilo para baixa presença
        self.estilo_baixa_presenca = NamedStyle(name="baixa_presenca")
        self.estilo_baixa_presenca.font = Font(size=10, color="000000")
        self.estilo_baixa_presenca.fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )
        self.estilo_baixa_presenca.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.estilo_baixa_presenca.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def gerar_consolidado_periodo(self, dados, nome_arquivo=None):
        """
        Gera relatório consolidado por período (grau).

        Args:
            dados: Dados processados pelo service
            nome_arquivo: Nome do arquivo (opcional)

        Returns:
            BytesIO: Arquivo Excel em memória
        """
        try:
            # Carregar template se disponível
            if self.template_path and os.path.exists(self.template_path):
                wb = load_workbook(self.template_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Consolidado Período"

            # Configurar título
            self._configurar_titulo_consolidado(ws, dados)

            # Configurar cabeçalhos
            self._configurar_cabecalhos_consolidado(ws, dados["periodo"]["meses"])

            # Preencher dados dos alunos
            self._preencher_dados_consolidado(
                ws, dados["alunos"], dados["periodo"]["meses"]
            )

            # Aplicar formatação
            self._aplicar_formatacao_consolidado(ws, dados)

            # Ajustar larguras das colunas
            self._ajustar_larguras_colunas(ws)

            # Salvar em memória
            arquivo_memoria = io.BytesIO()
            wb.save(arquivo_memoria)
            arquivo_memoria.seek(0)

            logger.info("Relatório consolidado gerado com sucesso")
            return arquivo_memoria

        except Exception as e:
            logger.error(f"Erro ao gerar relatório consolidado: {e}")
            raise

    def _configurar_titulo_consolidado(self, ws, dados):
        """Configura título do relatório consolidado."""
        # Título principal
        ws["A1"] = f"Relatório Consolidado de Presença - {dados['turma']['nome']}"
        ws["A1"].style = self.estilo_titulo

        # Período
        periodo_texto = f"Período: {dados['periodo']['inicio'].strftime('%d/%m/%Y')} a {dados['periodo']['fim'].strftime('%d/%m/%Y')}"
        ws["A2"] = periodo_texto
        ws["A2"].font = Font(size=12, italic=True)

        # Curso
        ws["A3"] = f"Curso: {dados['turma']['curso']}"
        ws["A3"].font = Font(size=11)

    def _configurar_cabecalhos_consolidado(self, ws, meses):
        """Configura cabeçalhos do relatório consolidado."""
        # Cabeçalhos fixos
        cabecalhos_fixos = ["Nº Iniciático", "Nome do Aluno", "Situação"]

        # Cabeçalhos dos meses
        cabecalhos_meses = []
        for mes in meses:
            cabecalhos_meses.extend(
                [
                    f"{mes['nome']}/{mes['ano']} P",
                    f"{mes['nome']}/{mes['ano']} F",
                    f"{mes['nome']}/{mes['ano']} J",
                ]
            )

        # Cabeçalhos de totais
        cabecalhos_totais = ["Total P", "Total F", "Total J", "% Presença"]

        # Combinar todos os cabeçalhos
        todos_cabecalhos = cabecalhos_fixos + cabecalhos_meses + cabecalhos_totais

        # Aplicar cabeçalhos na linha 5
        for col, cabecalho in enumerate(todos_cabecalhos, 1):
            cell = ws.cell(row=5, column=col, value=cabecalho)
            cell.style = self.estilo_cabecalho

    def _preencher_dados_consolidado(self, ws, alunos, meses):
        """Preenche dados dos alunos no relatório consolidado."""
        row = 6  # Começar após cabeçalhos

        for aluno_dados in alunos:
            aluno = aluno_dados["aluno"]
            col = 1

            # Dados fixos do aluno
            ws.cell(
                row=row, column=col, value=aluno.numero_iniciatico or "N/A"
            ).style = self.estilo_dados
            col += 1
            ws.cell(row=row, column=col, value=aluno.nome).style = self.estilo_dados
            col += 1
            ws.cell(
                row=row, column=col, value=aluno.get_situacao_display()
            ).style = self.estilo_dados
            col += 1

            # Dados por mês
            for mes in meses:
                chave_mes = mes["chave"]
                dados_mes = aluno_dados["meses"].get(
                    chave_mes, {"P": 0, "F": 0, "J": 0}
                )

                # P, F, J para cada mês
                ws.cell(
                    row=row, column=col, value=dados_mes["P"]
                ).style = self.estilo_dados
                col += 1
                ws.cell(
                    row=row, column=col, value=dados_mes["F"]
                ).style = self.estilo_dados
                col += 1
                ws.cell(
                    row=row, column=col, value=dados_mes["J"]
                ).style = self.estilo_dados
                col += 1

            # Totais
            totais = aluno_dados["totais"]
            ws.cell(row=row, column=col, value=totais["P"]).style = self.estilo_dados
            col += 1
            ws.cell(row=row, column=col, value=totais["F"]).style = self.estilo_dados
            col += 1
            ws.cell(row=row, column=col, value=totais["J"]).style = self.estilo_dados
            col += 1

            # Percentual de presença
            percentual_cell = ws.cell(
                row=row, column=col, value=f"{aluno_dados['percentual_presenca']}%"
            )

            # Aplicar formatação condicional para baixa presença
            if aluno_dados["percentual_presenca"] < 75:
                percentual_cell.style = self.estilo_baixa_presenca
                # Aplicar estilo de baixa presença para toda a linha
                for c in range(1, col + 1):
                    ws.cell(row=row, column=c).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )
            else:
                percentual_cell.style = self.estilo_dados

            row += 1

    def gerar_apuracao_mensal(self, dados, nome_arquivo=None):
        """
        Gera relatório de apuração mensal (mes01-99).

        Args:
            dados: Dados processados pelo service
            nome_arquivo: Nome do arquivo (opcional)

        Returns:
            BytesIO: Arquivo Excel em memória
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Mes{dados['periodo']['mes']:02d}_{dados['periodo']['ano']}"

            # Configurar título
            ws["A1"] = f"Apuração Mensal - {dados['turma']['nome']}"
            ws["A1"].style = self.estilo_titulo

            ws["A2"] = f"{dados['periodo']['nome_mes']}/{dados['periodo']['ano']}"
            ws["A2"].font = Font(size=12, italic=True)

            # Configurar cabeçalhos
            cabecalhos = ["Nº Iniciático", "Nome do Aluno"]

            # Adicionar dias do mês
            for dia_info in dados["periodo"]["dias_mes"]:
                cabecalhos.append(f"Dia {dia_info['dia']}")

            # Adicionar totais
            cabecalhos.extend(["Total P", "Total F", "Total J", "% Presença"])

            # Aplicar cabeçalhos
            for col, cabecalho in enumerate(cabecalhos, 1):
                cell = ws.cell(row=4, column=col, value=cabecalho)
                cell.style = self.estilo_cabecalho

            # Preencher dados
            row = 5
            for aluno_dados in dados["alunos"]:
                aluno = aluno_dados["aluno"]
                col = 1

                # Dados do aluno
                ws.cell(
                    row=row, column=col, value=aluno.numero_iniciatico or "N/A"
                ).style = self.estilo_dados
                col += 1
                ws.cell(row=row, column=col, value=aluno.nome).style = self.estilo_dados
                col += 1

                # Status por dia
                for dia_info in dados["periodo"]["dias_mes"]:
                    dia = dia_info["dia"]
                    status = ""

                    if dia in aluno_dados["dias"]:
                        status = aluno_dados["dias"][dia]["status"]

                    ws.cell(row=row, column=col, value=status).style = self.estilo_dados
                    col += 1

                # Totais
                totais = aluno_dados["totais"]
                ws.cell(
                    row=row, column=col, value=totais["P"]
                ).style = self.estilo_dados
                col += 1
                ws.cell(
                    row=row, column=col, value=totais["F"]
                ).style = self.estilo_dados
                col += 1
                ws.cell(
                    row=row, column=col, value=totais["J"]
                ).style = self.estilo_dados
                col += 1

                # Percentual
                total_atividades = sum(totais.values())
                percentual = (
                    (totais["P"] / total_atividades * 100)
                    if total_atividades > 0
                    else 0
                )
                ws.cell(
                    row=row, column=col, value=f"{percentual:.1f}%"
                ).style = self.estilo_dados

                row += 1

            # Ajustar larguras
            self._ajustar_larguras_colunas(ws)

            # Salvar em memória
            arquivo_memoria = io.BytesIO()
            wb.save(arquivo_memoria)
            arquivo_memoria.seek(0)

            return arquivo_memoria

        except Exception as e:
            logger.error(f"Erro ao gerar apuração mensal: {e}")
            raise

    def gerar_formulario_coleta(self, dados, nome_arquivo=None):
        """
        Gera formulário de coleta mensal (mod).

        Args:
            dados: Dados processados pelo service
            nome_arquivo: Nome do arquivo (opcional)

        Returns:
            BytesIO: Arquivo Excel em memória
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Coleta_{dados['periodo']['mes']:02d}_{dados['periodo']['ano']}"

            # Título
            ws["A1"] = f"Formulário de Coleta - {dados['turma']['nome']}"
            ws["A1"].style = self.estilo_titulo

            ws["A2"] = f"{dados['periodo']['nome_mes']}/{dados['periodo']['ano']}"
            ws["A2"].font = Font(size=12, italic=True)

            # Instruções
            ws["A3"] = (
                "Instruções: P=Presente, F=Falta, J=Justificada, V1=Voluntário Extra, V2=Voluntário Simples"
            )
            ws["A3"].font = Font(size=10, italic=True)

            # Cabeçalhos
            cabecalhos = ["Nº Iniciático", "Nome do Aluno"]

            for dia_info in dados["periodo"]["dias_mes"]:
                cabecalhos.append(f"{dia_info['dia']}")

            cabecalhos.append("Observações")

            # Aplicar cabeçalhos
            for col, cabecalho in enumerate(cabecalhos, 1):
                cell = ws.cell(row=5, column=col, value=cabecalho)
                cell.style = self.estilo_cabecalho

            # Preencher estrutura para preenchimento manual
            row = 6
            for aluno in dados["alunos"]:
                col = 1

                # Dados do aluno
                ws.cell(
                    row=row, column=col, value=aluno.numero_iniciatico or "N/A"
                ).style = self.estilo_dados
                col += 1
                ws.cell(row=row, column=col, value=aluno.nome).style = self.estilo_dados
                col += 1

                # Células vazias para preenchimento
                for dia_info in dados["periodo"]["dias_mes"]:
                    cell = ws.cell(row=row, column=col, value="")
                    cell.style = self.estilo_dados
                    # Destacar fins de semana
                    if dia_info["eh_fim_semana"]:
                        cell.fill = PatternFill(
                            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
                        )
                    col += 1

                # Célula de observações
                ws.cell(row=row, column=col, value="").style = self.estilo_dados

                row += 1

            # Ajustar larguras
            self._ajustar_larguras_colunas(ws)

            # Salvar em memória
            arquivo_memoria = io.BytesIO()
            wb.save(arquivo_memoria)
            arquivo_memoria.seek(0)

            return arquivo_memoria

        except Exception as e:
            logger.error(f"Erro ao gerar formulário coleta: {e}")
            raise

    def gerar_controle_geral(self, dados, nome_arquivo=None):
        """
        Gera relatório de controle geral da turma (pcg).

        Args:
            dados: Dados processados pelo service
            nome_arquivo: Nome do arquivo (opcional)

        Returns:
            BytesIO: Arquivo Excel em memória
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Controle_Geral"

            # Título
            ws["A1"] = "RELATÓRIO DE CONTROLE GERAL DA TURMA"
            ws["A1"].style = self.estilo_titulo

            # Dados da turma
            row = 3
            dados_turma = dados["turma"]

            campos_turma = [
                ("Nome da Turma:", dados_turma["nome"]),
                ("Curso:", dados_turma["curso"]),
                ("Descrição:", dados_turma.get("descricao", "N/A")),
                ("Nº do Livro:", dados_turma.get("num_livro", "N/A")),
                (
                    "% Mínimo de Presença:",
                    dados_turma.get("perc_presenca_minima", "N/A"),
                ),
                ("Data de Iniciação:", dados_turma.get("data_iniciacao", "N/A")),
                ("Data Início Atividades:", dados_turma.get("data_inicio_ativ", "N/A")),
                ("Status:", dados_turma.get("status", "N/A")),
                ("Instrutor:", dados_turma.get("instrutor", "N/A")),
            ]

            for campo, valor in campos_turma:
                ws.cell(row=row, column=1, value=campo).font = Font(bold=True)
                ws.cell(row=row, column=2, value=valor)
                row += 1

            # Estatísticas
            row += 2
            ws.cell(
                row=row, column=1, value="ESTATÍSTICAS DA TURMA"
            ).style = self.estilo_cabecalho
            row += 1

            estatisticas = dados.get("estatisticas", {})
            campos_stats = [
                ("Total de Alunos:", estatisticas.get("total_alunos", 0)),
                ("Alunos Ativos:", estatisticas.get("alunos_ativos", 0)),
                ("Vagas Disponíveis:", estatisticas.get("vagas_disponiveis", 0)),
                ("% Ocupação:", f"{estatisticas.get('percentual_ocupacao', 0)}%"),
            ]

            for campo, valor in campos_stats:
                ws.cell(row=row, column=1, value=campo).font = Font(bold=True)
                ws.cell(row=row, column=2, value=valor)
                row += 1

            # Lista de alunos
            row += 2
            ws.cell(
                row=row, column=1, value="ALUNOS MATRICULADOS"
            ).style = self.estilo_cabecalho
            row += 1

            # Cabeçalhos da lista
            cabecalhos_alunos = [
                "Nº Iniciático",
                "Nome",
                "Situação",
                "E-mail",
                "Telefone",
            ]
            for col, cabecalho in enumerate(cabecalhos_alunos, 1):
                ws.cell(
                    row=row, column=col, value=cabecalho
                ).style = self.estilo_cabecalho
            row += 1

            # Dados dos alunos
            for aluno in dados.get("alunos", []):
                ws.cell(
                    row=row, column=1, value=aluno.numero_iniciatico or "N/A"
                ).style = self.estilo_dados
                ws.cell(row=row, column=2, value=aluno.nome).style = self.estilo_dados
                ws.cell(
                    row=row, column=3, value=aluno.get_situacao_display()
                ).style = self.estilo_dados
                ws.cell(row=row, column=4, value=aluno.email).style = self.estilo_dados
                ws.cell(
                    row=row, column=5, value=getattr(aluno, "celular", "N/A")
                ).style = self.estilo_dados
                row += 1

            # Rodapé
            row += 2
            ws.cell(
                row=row,
                column=1,
                value=f"Relatório gerado em: {dados['data_geracao'].strftime('%d/%m/%Y %H:%M')}",
            )

            # Ajustar larguras
            self._ajustar_larguras_colunas(ws)

            # Salvar em memória
            arquivo_memoria = io.BytesIO()
            wb.save(arquivo_memoria)
            arquivo_memoria.seek(0)

            return arquivo_memoria

        except Exception as e:
            logger.error(f"Erro ao gerar controle geral: {e}")
            raise

    def _aplicar_formatacao_consolidado(self, ws, dados):
        """Aplica formatação específica do relatório consolidado."""
        # Mesclar células do título
        ws.merge_cells("A1:F1")
        ws.merge_cells("A2:F2")
        ws.merge_cells("A3:F3")

        # Aplicar bordas
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

    def _ajustar_larguras_colunas(self, ws):
        """Ajusta larguras das colunas automaticamente."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _calcular_percentual_presenca(self, presencas, total):
        """Calcula percentual de presença."""
        return round((presencas / total) * 100, 2) if total > 0 else 0
