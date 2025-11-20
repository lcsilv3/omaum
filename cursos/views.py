import csv
import unicodedata
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import HttpResponse
from django.shortcuts import redirect, render
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from . import services
from .forms import CursoForm
from .models import Curso


def _normalizar_coluna(chave: Any) -> str:
    """Normaliza o nome da coluna removendo acentos, espaços e underscores."""

    texto = services.normalizar_texto(chave)
    base = (
        unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    )
    return base.replace(" ", "").replace("_", "").lower()


def _mapear_registro(row: Dict[str, Any], linha_planilha: int) -> Dict[str, Any]:
    """Transforma uma linha da planilha em dicionário compatível com o serviço."""

    colunas_normalizadas = {
        _normalizar_coluna(coluna): valor for coluna, valor in row.items() if coluna
    }

    return {
        "linha": linha_planilha,
        "id": colunas_normalizadas.get("id") or colunas_normalizadas.get("codigo"),
        "nome": colunas_normalizadas.get("nome")
        or colunas_normalizadas.get("curso")
        or colunas_normalizadas.get("titulo"),
        "descricao": colunas_normalizadas.get("descricao")
        or colunas_normalizadas.get("descricaodocurso"),
        "ativo": colunas_normalizadas.get("ativo")
        or colunas_normalizadas.get("status"),
    }


def _linha_vazia(row: Dict[str, Any]) -> bool:
    """Retorna True quando todas as células da linha estão vazias."""

    return not any(services.normalizar_texto(valor) for valor in row.values())


def _ler_csv(arquivo) -> List[Dict[str, Any]]:
    """Lê arquivo CSV e devolve lista normalizada para sincronização."""

    arquivo.seek(0)
    conteudo = arquivo.read()
    if isinstance(conteudo, bytes):
        texto = conteudo.decode("utf-8-sig")
    else:
        texto = conteudo

    buffer = StringIO(texto)
    amostra = buffer.read(2048)
    buffer.seek(0)
    delimitador = ";" if amostra.count(";") > amostra.count(",") else ","

    reader = csv.DictReader(buffer, delimiter=delimitador)
    registros: List[Dict[str, Any]] = []
    for indice, row in enumerate(reader, start=2):
        if _linha_vazia(row):
            continue
        registros.append(_mapear_registro(row, indice))

    arquivo.seek(0)
    return registros


def _ler_xlsx(arquivo) -> List[Dict[str, Any]]:
    """Lê arquivo XLSX e devolve lista normalizada para sincronização."""

    arquivo.seek(0)
    workbook = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    sheet = workbook.active
    headers: List[str] = []
    registros: List[Dict[str, Any]] = []

    for indice, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        valores = [services.normalizar_texto(valor) for valor in row]
        if indice == 1:
            headers = valores
            continue
        row_dict: Dict[str, Any] = {
            headers[posicao] if posicao < len(headers) else f"coluna_{posicao}": valor
            for posicao, valor in enumerate(valores)
        }
        if _linha_vazia(row_dict):
            continue
        registros.append(_mapear_registro(row_dict, indice))

    workbook.close()
    return registros


@login_required
def listar_cursos(request):
    """Lista todos os cursos cadastrados, com filtro de busca e suporte a AJAX."""
    query = request.GET.get("q", "")
    cursos = services.listar_cursos(query=query)
    context = {"cursos": cursos, "query": query}

    # Se for uma requisição HTMX/AJAX, retorna apenas o fragmento da tabela
    if request.headers.get("HX-Request") == "true":
        return render(request, "cursos/partials/tabela_cursos.html", context)

    # Caso contrário, renderiza a página completa
    context.update(
        {
            "pagamentos_atrasados": [],
            "pagamentos_atrasados_count": 0,
        }
    )
    return render(request, "cursos/listar_cursos.html", context)


@login_required
def relatorio_cursos(request):
    """Gera o relatório de cursos em HTML ou exporta em CSV, Excel ou PDF."""
    formato = request.GET.get("formato")
    cursos = services.listar_cursos()

    if formato == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="relatorio_cursos.csv"'
        writer = csv.writer(response)
        writer.writerow(["ID", "Nome", "Descrição"])
        for curso in cursos:
            writer.writerow([curso.id, curso.nome, curso.descricao])
        return response

    if formato == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="relatorio_cursos.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Cursos"
        worksheet.append(["ID", "Nome", "Descrição"])
        for curso in cursos:
            worksheet.append([curso.id, curso.nome, curso.descricao])
        workbook.save(response)
        return response

    if formato == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="relatorio_cursos.pdf"'
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        data_emissao = datetime.now().strftime("%d/%m/%Y %H:%M")
        p.setFont("Helvetica-Bold", 12)
        p.drawCentredString(
            width / 2.0,
            height - inch,
            "OMAUM - Ordem Mística de Aspiração Universal ao Mestrado",
        )
        p.setFont("Helvetica", 10)
        p.drawCentredString(
            width / 2.0, height - 1.2 * inch, "Sistema de Gestão Integrada"
        )
        p.setFont("Helvetica", 8)
        p.drawRightString(
            width - inch, height - 1.2 * inch, f"Data de Emissão: {data_emissao}"
        )
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width / 2.0, height - 1.5 * inch, "Relatório de Cursos")
        y = height - 1.9 * inch
        p.setFont("Helvetica-Bold", 12)
        p.drawString(inch, y, "ID")
        p.drawString(1.5 * inch, y, "Nome")
        p.drawString(4 * inch, y, "Descrição")
        p.line(inch, y - 0.1 * inch, width - inch, y - 0.1 * inch)
        y -= 0.3 * inch
        p.setFont("Helvetica", 10)
        for curso in cursos:
            if y < inch:
                p.showPage()
                p.setFont("Helvetica-Bold", 12)
                p.drawCentredString(
                    width / 2.0, height - inch, "Relatório de Cursos (Continuação)"
                )
                y = height - 1.5 * inch
                p.setFont("Helvetica-Bold", 12)
                p.drawString(inch, y, "ID")
                p.drawString(1.5 * inch, y, "Nome")
                p.drawString(4 * inch, y, "Descrição")
                p.line(inch, y - 0.1 * inch, width - inch, y - 0.1 * inch)
                y -= 0.3 * inch
                p.setFont("Helvetica", 10)
            p.drawString(inch, y, str(curso.id))
            p.drawString(1.5 * inch, y, curso.nome)
            descricao_text = curso.descricao if curso.descricao else ""
            text_object = p.beginText(4 * inch, y)
            text_object.setFont("Helvetica", 10)
            wrapped_text = [
                descricao_text[i : i + 50] for i in range(0, len(descricao_text), 50)
            ]
            for line in wrapped_text:
                text_object.textLine(line)
            p.drawText(text_object)
            y -= (0.2 * inch * len(wrapped_text)) + 0.1 * inch
        p.showPage()
        p.save()
        return response

    # Renderiza o HTML por padrão
    data_emissao = datetime.now()
    context = {
        "titulo": "Relatório de Cursos",
        "cursos": cursos,
        "data_emissao": data_emissao.strftime("%d/%m/%Y %H:%M"),
        "nome_organizacao": "OMAUM - Ordem Mística de Aspiração Universal ao Mestrado",
        "nome_sistema": "Sistema de Gestão Integrada",
    }
    return render(request, "cursos/relatorio_cursos.html", context)


@login_required
def criar_curso(request):
    """Cria um novo curso utilizando a camada de serviço."""
    if request.method == "POST":
        form = CursoForm(request.POST)
        if form.is_valid():
            try:
                services.criar_curso(
                    nome=form.cleaned_data["nome"],
                    descricao=form.cleaned_data["descricao"],
                )
                messages.success(request, "Curso criado com sucesso!")
                return redirect("cursos:listar_cursos")
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = CursoForm()
    return render(request, "cursos/criar_curso.html", {"form": form})


@login_required
def detalhar_curso(request, id):
    """Exibe os detalhes de um curso."""
    curso = services.obter_curso_por_id(id)
    if not curso:
        messages.error(request, "Curso não encontrado.")
        return redirect("cursos:listar_cursos")
    return render(request, "cursos/detalhar_curso.html", {"curso": curso})


@login_required
def editar_curso(request, id):
    """Edita um curso existente utilizando a camada de serviço."""
    curso = services.obter_curso_por_id(id)
    if not curso:
        messages.error(request, "Curso não encontrado.")
        return redirect("cursos:listar_cursos")

    if request.method == "POST":
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            services.atualizar_curso(
                curso_id=id,
                nome=form.cleaned_data["nome"],
                descricao=form.cleaned_data["descricao"],
            )
            messages.success(request, "Curso atualizado com sucesso!")
            return redirect("cursos:listar_cursos")
    else:
        form = CursoForm(instance=curso)
    return render(request, "cursos/editar_curso.html", {"form": form, "curso": curso})


@login_required
def excluir_curso(request, id):
    """Exclui um curso utilizando a camada de serviço."""
    curso = services.obter_curso_por_id(id)
    if not curso:
        messages.error(request, "Curso não encontrado.")
        return redirect("cursos:listar_cursos")

    dependencias = services.verificar_dependencias_curso(curso)

    if request.method == "POST":
        try:
            if services.excluir_curso(id):
                messages.success(request, "Curso excluído com sucesso!")
                return redirect("cursos:listar_cursos")
            else:
                messages.error(request, "Não foi possível excluir o curso.")
        except ValueError as e:
            messages.error(request, str(e))
            return redirect("cursos:excluir_curso", id=id)
        except IntegrityError:
            messages.error(request, "Erro de integridade. Verifique as dependências.")
            return redirect("cursos:excluir_curso", id=id)

    return render(
        request,
        "cursos/excluir_curso.html",
        {"curso": curso, "dependencias": dependencias},
    )


@login_required
def importar_cursos(request):
    """Importa cursos a partir de planilha CSV ou XLSX utilizando a camada de serviço."""
    if request.method == "POST" and request.FILES.get("arquivo"):
        arquivo = request.FILES["arquivo"]
        extensao = Path(arquivo.name).suffix.lower()

        try:
            registros: List[Dict[str, Any]]
            if extensao in {".csv"}:
                registros = _ler_csv(arquivo)
            elif extensao in {".xlsx", ".xlsm"}:
                registros = _ler_xlsx(arquivo)
            else:
                messages.error(
                    request,
                    "Formato não suportado. Utilize arquivos CSV ou XLSX.",
                )
                return redirect("cursos:importar_cursos")

            if not registros:
                messages.warning(
                    request, "Nenhum registro válido encontrado no arquivo."
                )
                return redirect("cursos:importar_cursos")

            resumo = services.sincronizar_cursos(
                registros,
                desativar_nao_listados=False,
            )

            if resumo.get("processados", 0):
                mensagem_sucesso = (
                    f"Importação concluída: {resumo['processados']} processados, "
                    f"{resumo['criados']} criados, {resumo['atualizados']} atualizados, "
                    f"{resumo['reativados']} reativados."
                )
                messages.success(request, mensagem_sucesso)
            else:
                messages.warning(
                    request,
                    "Nenhum curso foi processado. Verifique o conteúdo da planilha.",
                )

            avisos = resumo.get("avisos", [])
            if avisos:
                messages.warning(
                    request,
                    f"Foram encontrados {len(avisos)} avisos durante a importação.",
                )
                for aviso in avisos[:5]:
                    messages.info(request, aviso)
                if len(avisos) > 5:
                    messages.info(
                        request,
                        f"... e mais {len(avisos) - 5} avisos não exibidos.",
                    )

            return redirect("cursos:listar_cursos")
        except Exception as exc:  # pragma: no cover - proteção contra erros imprevistos
            messages.error(request, f"Erro ao importar cursos: {exc}")
            return redirect("cursos:importar_cursos")

    return render(request, "cursos/importar_cursos.html")
