"""
Views para exportação avançada de dados de presenças.
Implementa múltiplos formatos, formatação profissional e agendamento.
"""

import os
import io
import csv
import logging
import tempfile
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from io import StringIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import TemplateView
from django.core.mail import EmailMessage
from django.conf import settings
from django.utils import timezone
from django.db.models import Q, Sum, Avg, Count
from django.core.management.base import BaseCommand
from django.template.loader import render_to_string

# Verificar se as bibliotecas estão disponíveis
try:
    import openpyxl
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference, Series
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from celery import shared_task
    CELERY_AVAILABLE = True
except ImportError:
    CELERY_AVAILABLE = False
    # Mock shared_task se Celery não estiver disponível
    def shared_task(func):
        return func

from ..models import (
    PresencaDetalhada, Aluno, Turma
)
from .consolidado import ConsolidadoPresencasView


logger = logging.getLogger(__name__)


class ExportacaoAvancadaView(LoginRequiredMixin, TemplateView):
    """
    View principal para exportação avançada de dados.
    Interface para seleção de formato e configurações.
    """
    template_name = 'presencas/exportacao/exportacao_avancada.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Formatos disponíveis
        context['formatos'] = [
            {'valor': 'excel_basico', 'nome': 'Excel Básico (.xlsx)'},
            {'valor': 'excel_avancado', 'nome': 'Excel Profissional (.xlsx)'},
            {'valor': 'excel_graficos', 'nome': 'Excel com Gráficos (.xlsx)'},
            {'valor': 'csv', 'nome': 'CSV (.csv)'},
            {'valor': 'pdf_simples', 'nome': 'PDF Simples (.pdf)'},
            {'valor': 'pdf_completo', 'nome': 'PDF Completo (.pdf)'},
        ]
        
        # Templates de relatório
        context['templates'] = [
            {'valor': 'consolidado_geral', 'nome': 'Consolidado Geral'},
            {'valor': 'por_turma', 'nome': 'Relatório por Turma'},
            {'valor': 'por_curso', 'nome': 'Relatório por Curso'},
            {'valor': 'estatisticas_executivas', 'nome': 'Estatísticas Executivas'},
            {'valor': 'carencia_presencas', 'nome': 'Relatório de Carência'},
            {'valor': 'comparativo_temporal', 'nome': 'Comparativo Temporal'},
        ]
        
        # Opções de período
        context['periodos'] = [
            {'valor': 'atual', 'nome': 'Período Atual'},
            {'valor': 'ultimo_mes', 'nome': 'Último Mês'},
            {'valor': 'ultimo_trimestre', 'nome': 'Último Trimestre'},
            {'valor': 'ultimo_semestre', 'nome': 'Último Semestre'},
            {'valor': 'ano_atual', 'nome': 'Ano Atual'},
            {'valor': 'personalizado', 'nome': 'Período Personalizado'},
        ]
        
        # Filtros disponíveis
        context['turmas'] = Turma.objects.filter(ativa=True).order_by('nome')
        context['cursos'] = Turma.objects.filter(ativa=True).values_list(
            'curso__nome', flat=True
        ).distinct().order_by('curso__nome')
        
        return context


class ProcessarExportacaoView(LoginRequiredMixin, View):
    """
    View para processar solicitações de exportação.
    """
    
    def post(self, request, *args, **kwargs):
        """Processa requisição de exportação."""
        try:
            formato = request.POST.get('formato')
            template = request.POST.get('template')
            configuracoes = self._extrair_configuracoes(request)
            
            # Validar parâmetros
            if not formato or not template:
                return JsonResponse({
                    'erro': 'Formato e template são obrigatórios'
                }, status=400)
            
            # Processar exportação
            resultado = self._processar_exportacao(formato, template, configuracoes)
            
            if 'erro' in resultado:
                return JsonResponse(resultado, status=500)
            
            return resultado
            
        except Exception as e:
            logger.error(f"Erro ao processar exportação: {str(e)}")
            return JsonResponse({
                'erro': f'Erro interno: {str(e)}'
            }, status=500)
    
    def _extrair_configuracoes(self, request) -> Dict[str, Any]:
        """Extrai configurações da requisição."""
        return {
            'periodo': request.POST.get('periodo', 'atual'),
            'data_inicio': request.POST.get('data_inicio'),
            'data_fim': request.POST.get('data_fim'),
            'turma_id': request.POST.get('turma_id'),
            'curso': request.POST.get('curso'),
            'incluir_graficos': request.POST.get('incluir_graficos') == 'on',
            'incluir_estatisticas': request.POST.get('incluir_estatisticas') == 'on',
            'formatacao_profissional': request.POST.get('formatacao_profissional') == 'on',
            'multiplas_abas': request.POST.get('multiplas_abas') == 'on',
            'email_envio': request.POST.get('email_envio'),
            'titulo_personalizado': request.POST.get('titulo_personalizado', ''),
        }
    
    def _processar_exportacao(self, formato: str, template: str, config: Dict[str, Any]):
        """Processa a exportação conforme formato e template."""
        
        # Obter dados
        dados = self._obter_dados(template, config)
        
        # Gerar arquivo conforme formato
        if formato.startswith('excel'):
            return self._gerar_excel(formato, template, dados, config)
        elif formato == 'csv':
            return self._gerar_csv(template, dados, config)
        elif formato.startswith('pdf'):
            return self._gerar_pdf(formato, template, dados, config)
        else:
            return {'erro': f'Formato não suportado: {formato}'}
    
    def _obter_dados(self, template: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados conforme template e configurações."""
        
        # Criar instância do consolidado para reaproveitar lógica
        consolidado_view = ConsolidadoPresencasView()
        consolidado_view.request = self.request
        
        # Aplicar filtros baseados nas configurações
        filtros = self._construir_filtros(config)
        
        if template == 'consolidado_geral':
            return self._obter_dados_consolidado_geral(filtros)
        elif template == 'por_turma':
            return self._obter_dados_por_turma(filtros)
        elif template == 'por_curso':
            return self._obter_dados_por_curso(filtros)
        elif template == 'estatisticas_executivas':
            return self._obter_estatisticas_executivas(filtros)
        elif template == 'carencia_presencas':
            return self._obter_dados_carencia(filtros)
        elif template == 'comparativo_temporal':
            return self._obter_dados_comparativo(filtros)
        else:
            return {'erro': f'Template não suportado: {template}'}
    
    def _construir_filtros(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Constrói filtros a partir das configurações."""
        filtros = {}
        
        # Filtros de período
        if config.get('periodo') != 'personalizado':
            if config['periodo'] == 'ultimo_mes':
                filtros['data_inicio'] = timezone.now() - timedelta(days=30)
            elif config['periodo'] == 'ultimo_trimestre':
                filtros['data_inicio'] = timezone.now() - timedelta(days=90)
            elif config['periodo'] == 'ultimo_semestre':
                filtros['data_inicio'] = timezone.now() - timedelta(days=180)
            elif config['periodo'] == 'ano_atual':
                filtros['data_inicio'] = timezone.now().replace(month=1, day=1)
        else:
            if config.get('data_inicio'):
                filtros['data_inicio'] = datetime.strptime(
                    config['data_inicio'], '%Y-%m-%d'
                ).date()
            if config.get('data_fim'):
                filtros['data_fim'] = datetime.strptime(
                    config['data_fim'], '%Y-%m-%d'
                ).date()
        
        # Filtros de turma e curso
        if config.get('turma_id'):
            filtros['turma_id'] = config['turma_id']
        
        if config.get('curso'):
            filtros['curso'] = config['curso']
        
        return filtros
    
    def _obter_dados_consolidado_geral(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados para relatório consolidado geral."""
        
        # Reutilizar lógica do consolidado existente
        consolidado_view = ConsolidadoPresencasView()
        consolidado_view.request = self.request
        
        presencas_detalhadas = consolidado_view.obter_presencas_consolidadas(filtros)
        atividades = consolidado_view.obter_atividades_filtradas(filtros)
        
        # Calcular estatísticas gerais
        estatisticas = self._calcular_estatisticas_gerais(presencas_detalhadas)
        
        return {
            'presencas_detalhadas': presencas_detalhadas,
            'atividades': atividades,
            'estatisticas': estatisticas,
            'filtros_aplicados': filtros,
        }
    
    def _obter_dados_por_turma(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados agrupados por turma."""
        
        # Base query
        query = PresencaDetalhada.objects.select_related(
            'aluno', 'turma', 'atividade'
        )
        
        # Aplicar filtros
        if filtros.get('data_inicio'):
            query = query.filter(atividade__data__gte=filtros['data_inicio'])
        if filtros.get('data_fim'):
            query = query.filter(atividade__data__lte=filtros['data_fim'])
        if filtros.get('turma_id'):
            query = query.filter(turma_id=filtros['turma_id'])
        
        # Agrupar por turma
        dados_por_turma = {}
        for presenca in query:
            turma_id = presenca.turma.id
            if turma_id not in dados_por_turma:
                dados_por_turma[turma_id] = {
                    'turma': presenca.turma,
                    'presencas': [],
                    'estatisticas': {
                        'total_alunos': set(),
                        'total_presencas': 0,
                        'total_faltas': 0,
                        'percentual_medio': 0,
                    }
                }
            
            dados_por_turma[turma_id]['presencas'].append(presenca)
            dados_por_turma[turma_id]['estatisticas']['total_alunos'].add(
                presenca.aluno.id
            )
            dados_por_turma[turma_id]['estatisticas']['total_presencas'] += presenca.presencas
            dados_por_turma[turma_id]['estatisticas']['total_faltas'] += presenca.faltas
        
        # Finalizar estatísticas
        for turma_data in dados_por_turma.values():
            stats = turma_data['estatisticas']
            stats['total_alunos'] = len(stats['total_alunos'])
            total_convocacoes = stats['total_presencas'] + stats['total_faltas']
            if total_convocacoes > 0:
                stats['percentual_medio'] = (
                    stats['total_presencas'] / total_convocacoes * 100
                )
        
        return {
            'dados_por_turma': dados_por_turma,
            'filtros_aplicados': filtros,
        }
    
    def _obter_dados_por_curso(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados agrupados por curso."""
        
        # Obter dados por turma primeiro
        dados_turma = self._obter_dados_por_turma(filtros)
        
        # Reagrupar por curso
        dados_por_curso = {}
        for turma_data in dados_turma['dados_por_turma'].values():
            curso_nome = turma_data['turma'].curso.nome if turma_data['turma'].curso else 'Sem Curso'
            
            if curso_nome not in dados_por_curso:
                dados_por_curso[curso_nome] = {
                    'curso': curso_nome,
                    'turmas': [],
                    'estatisticas_curso': {
                        'total_alunos': 0,
                        'total_presencas': 0,
                        'total_faltas': 0,
                        'percentual_medio': 0,
                    }
                }
            
            dados_por_curso[curso_nome]['turmas'].append(turma_data)
            
            # Somar estatísticas
            stats_curso = dados_por_curso[curso_nome]['estatisticas_curso']
            stats_turma = turma_data['estatisticas']
            
            stats_curso['total_alunos'] += stats_turma['total_alunos']
            stats_curso['total_presencas'] += stats_turma['total_presencas']
            stats_curso['total_faltas'] += stats_turma['total_faltas']
        
        # Calcular percentuais médios
        for curso_data in dados_por_curso.values():
            stats = curso_data['estatisticas_curso']
            total_convocacoes = stats['total_presencas'] + stats['total_faltas']
            if total_convocacoes > 0:
                stats['percentual_medio'] = (
                    stats['total_presencas'] / total_convocacoes * 100
                )
        
        return {
            'dados_por_curso': dados_por_curso,
            'filtros_aplicados': filtros,
        }
    
    def _obter_estatisticas_executivas(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém estatísticas executivas avançadas."""
        
        # Dados base
        query = PresencaDetalhada.objects.select_related('aluno', 'turma', 'atividade')
        
        # Aplicar filtros
        if filtros.get('data_inicio'):
            query = query.filter(atividade__data__gte=filtros['data_inicio'])
        if filtros.get('data_fim'):
            query = query.filter(atividade__data__lte=filtros['data_fim'])
        
        # Estatísticas gerais
        estatisticas = {
            'resumo_geral': {
                'total_alunos': query.values('aluno').distinct().count(),
                'total_atividades': query.values('atividade').distinct().count(),
                'total_presencas': query.aggregate(Sum('presencas'))['presencas__sum'] or 0,
                'total_faltas': query.aggregate(Sum('faltas'))['faltas__sum'] or 0,
                'percentual_geral': 0,
            },
            'por_periodo': self._calcular_tendencias_temporais(query),
            'top_alunos_presenca': self._obter_top_alunos_presenca(query),
            'alunos_carencia': self._obter_alunos_carencia(query),
            'distribuicao_turmas': self._obter_distribuicao_turmas(query),
            'atividades_mais_participativas': self._obter_atividades_participativas(query),
        }
        
        # Calcular percentual geral
        total_conv = estatisticas['resumo_geral']['total_presencas'] + \
                     estatisticas['resumo_geral']['total_faltas']
        if total_conv > 0:
            estatisticas['resumo_geral']['percentual_geral'] = (
                estatisticas['resumo_geral']['total_presencas'] / total_conv * 100
            )
        
        return {
            'estatisticas_executivas': estatisticas,
            'filtros_aplicados': filtros,
        }
    
    def _obter_dados_carencia(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados de alunos com carência de presença."""
        
        # Definir limite de carência (percentual abaixo de 75%)
        limite_carencia = 75.0
        
        # Obter dados consolidados
        dados_consolidado = self._obter_dados_consolidado_geral(filtros)
        presencas_detalhadas = dados_consolidado['presencas_detalhadas']
        
        # Agrupar por aluno e calcular percentuais
        alunos_carencia = []
        dados_por_aluno = {}
        
        for presenca in presencas_detalhadas:
            aluno_id = presenca.aluno.id
            if aluno_id not in dados_por_aluno:
                dados_por_aluno[aluno_id] = {
                    'aluno': presenca.aluno,
                    'turma': presenca.turma,
                    'total_convocacoes': 0,
                    'total_presencas': 0,
                    'total_faltas': 0,
                    'percentual': 0,
                    'atividades_detalhes': []
                }
            
            dados_aluno = dados_por_aluno[aluno_id]
            dados_aluno['total_convocacoes'] += presenca.convocacoes
            dados_aluno['total_presencas'] += presenca.presencas
            dados_aluno['total_faltas'] += presenca.faltas
            dados_aluno['atividades_detalhes'].append({
                'atividade': presenca.atividade,
                'presencas': presenca.presencas,
                'faltas': presenca.faltas,
                'percentual': float(presenca.percentual_presenca)
            })
        
        # Filtrar alunos com carência
        for dados_aluno in dados_por_aluno.values():
            if dados_aluno['total_convocacoes'] > 0:
                dados_aluno['percentual'] = (
                    dados_aluno['total_presencas'] / dados_aluno['total_convocacoes'] * 100
                )
                
                if dados_aluno['percentual'] < limite_carencia:
                    alunos_carencia.append(dados_aluno)
        
        # Ordenar por percentual (menor primeiro)
        alunos_carencia.sort(key=lambda x: x['percentual'])
        
        return {
            'alunos_carencia': alunos_carencia,
            'limite_carencia': limite_carencia,
            'total_alunos_carencia': len(alunos_carencia),
            'filtros_aplicados': filtros,
        }
    
    def _obter_dados_comparativo(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados para comparativo temporal."""
        
        # Definir períodos para comparação
        agora = timezone.now()
        periodos = [
            {
                'nome': 'Mês Atual',
                'inicio': agora.replace(day=1),
                'fim': agora
            },
            {
                'nome': 'Mês Anterior',
                'inicio': (agora.replace(day=1) - timedelta(days=1)).replace(day=1),
                'fim': agora.replace(day=1) - timedelta(days=1)
            },
            {
                'nome': 'Trimestre Atual',
                'inicio': agora - timedelta(days=90),
                'fim': agora
            },
            {
                'nome': 'Trimestre Anterior',
                'inicio': agora - timedelta(days=180),
                'fim': agora - timedelta(days=90)
            }
        ]
        
        dados_comparativo = []
        
        for periodo in periodos:
            # Aplicar filtros de período
            filtros_periodo = filtros.copy()
            filtros_periodo['data_inicio'] = periodo['inicio'].date()
            filtros_periodo['data_fim'] = periodo['fim'].date()
            
            # Obter dados do período
            dados_periodo = self._obter_dados_consolidado_geral(filtros_periodo)
            estatisticas = dados_periodo['estatisticas']
            
            dados_comparativo.append({
                'periodo': periodo['nome'],
                'data_inicio': periodo['inicio'].date(),
                'data_fim': periodo['fim'].date(),
                'estatisticas': estatisticas
            })
        
        return {
            'dados_comparativo': dados_comparativo,
            'filtros_aplicados': filtros,
        }
    
    def _calcular_estatisticas_gerais(self, presencas_detalhadas) -> Dict[str, Any]:
        """Calcula estatísticas gerais dos dados."""
        
        if not presencas_detalhadas:
            return {
                'total_alunos': 0,
                'total_atividades': 0,
                'total_presencas': 0,
                'total_faltas': 0,
                'percentual_medio': 0,
                'voluntarios_total': 0,
            }
        
        alunos_unicos = set()
        atividades_unicas = set()
        total_presencas = 0
        total_faltas = 0
        total_voluntarios = 0
        
        for presenca in presencas_detalhadas:
            alunos_unicos.add(presenca.aluno.id)
            atividades_unicas.add(presenca.atividade.id)
            total_presencas += presenca.presencas
            total_faltas += presenca.faltas
            total_voluntarios += presenca.total_voluntarios
        
        total_convocacoes = total_presencas + total_faltas
        percentual_medio = (
            (total_presencas / total_convocacoes * 100) if total_convocacoes > 0 else 0
        )
        
        return {
            'total_alunos': len(alunos_unicos),
            'total_atividades': len(atividades_unicas),
            'total_presencas': total_presencas,
            'total_faltas': total_faltas,
            'percentual_medio': round(percentual_medio, 2),
            'voluntarios_total': total_voluntarios,
        }
    
    def _calcular_tendencias_temporais(self, query) -> List[Dict[str, Any]]:
        """Calcula tendências temporais dos dados."""
        
        # Agrupar por mês
        dados_mensais = {}
        
        for presenca in query:
            mes_ano = presenca.atividade.data.strftime('%Y-%m')
            if mes_ano not in dados_mensais:
                dados_mensais[mes_ano] = {
                    'mes_ano': mes_ano,
                    'presencas': 0,
                    'faltas': 0,
                    'total_alunos': set(),
                    'percentual': 0
                }
            
            dados_mensais[mes_ano]['presencas'] += presenca.presencas
            dados_mensais[mes_ano]['faltas'] += presenca.faltas
            dados_mensais[mes_ano]['total_alunos'].add(presenca.aluno.id)
        
        # Calcular percentuais e converter sets
        tendencias = []
        for dados in dados_mensais.values():
            total_conv = dados['presencas'] + dados['faltas']
            dados['percentual'] = (
                (dados['presencas'] / total_conv * 100) if total_conv > 0 else 0
            )
            dados['total_alunos'] = len(dados['total_alunos'])
            tendencias.append(dados)
        
        # Ordenar por mês
        tendencias.sort(key=lambda x: x['mes_ano'])
        
        return tendencias
    
    def _obter_top_alunos_presenca(self, query, limite: int = 10) -> List[Dict[str, Any]]:
        """Obtém top alunos por presença."""
        
        alunos_presenca = {}
        
        for presenca in query:
            aluno_id = presenca.aluno.id
            if aluno_id not in alunos_presenca:
                alunos_presenca[aluno_id] = {
                    'aluno': presenca.aluno,
                    'turma': presenca.turma,
                    'total_presencas': 0,
                    'total_faltas': 0,
                    'percentual': 0
                }
            
            alunos_presenca[aluno_id]['total_presencas'] += presenca.presencas
            alunos_presenca[aluno_id]['total_faltas'] += presenca.faltas
        
        # Calcular percentuais
        for dados in alunos_presenca.values():
            total_conv = dados['total_presencas'] + dados['total_faltas']
            dados['percentual'] = (
                (dados['total_presencas'] / total_conv * 100) if total_conv > 0 else 0
            )
        
        # Ordenar e limitar
        top_alunos = sorted(
            alunos_presenca.values(),
            key=lambda x: x['percentual'],
            reverse=True
        )[:limite]
        
        return top_alunos
    
    def _obter_alunos_carencia(self, query, limite: int = 10) -> List[Dict[str, Any]]:
        """Obtém alunos com maior carência de presença."""
        
        top_alunos = self._obter_top_alunos_presenca(query, limite=1000)
        
        # Filtrar e ordenar por menor percentual
        alunos_carencia = [
            aluno for aluno in top_alunos if aluno['percentual'] < 75
        ]
        
        return alunos_carencia[:limite]
    
    def _obter_distribuicao_turmas(self, query) -> List[Dict[str, Any]]:
        """Obtém distribuição de presenças por turma."""
        
        distribuicao = {}
        
        for presenca in query:
            turma_nome = presenca.turma.nome
            if turma_nome not in distribuicao:
                distribuicao[turma_nome] = {
                    'turma': turma_nome,
                    'total_presencas': 0,
                    'total_faltas': 0,
                    'total_alunos': set(),
                    'percentual': 0
                }
            
            distribuicao[turma_nome]['total_presencas'] += presenca.presencas
            distribuicao[turma_nome]['total_faltas'] += presenca.faltas
            distribuicao[turma_nome]['total_alunos'].add(presenca.aluno.id)
        
        # Calcular percentuais e converter
        resultado = []
        for dados in distribuicao.values():
            total_conv = dados['total_presencas'] + dados['total_faltas']
            dados['percentual'] = (
                (dados['total_presencas'] / total_conv * 100) if total_conv > 0 else 0
            )
            dados['total_alunos'] = len(dados['total_alunos'])
            resultado.append(dados)
        
        # Ordenar por percentual
        resultado.sort(key=lambda x: x['percentual'], reverse=True)
        
        return resultado
    
    def _obter_atividades_participativas(self, query) -> List[Dict[str, Any]]:
        """Obtém atividades com maior participação."""
        
        atividades = {}
        
        for presenca in query:
            atividade_id = presenca.atividade.id
            if atividade_id not in atividades:
                atividades[atividade_id] = {
                    'atividade': presenca.atividade,
                    'total_presencas': 0,
                    'total_faltas': 0,
                    'total_participantes': set(),
                    'percentual': 0
                }
            
            atividades[atividade_id]['total_presencas'] += presenca.presencas
            atividades[atividade_id]['total_faltas'] += presenca.faltas
            atividades[atividade_id]['total_participantes'].add(presenca.aluno.id)
        
        # Calcular percentuais e converter
        resultado = []
        for dados in atividades.values():
            total_conv = dados['total_presencas'] + dados['total_faltas']
            dados['percentual'] = (
                (dados['total_presencas'] / total_conv * 100) if total_conv > 0 else 0
            )
            dados['total_participantes'] = len(dados['total_participantes'])
            resultado.append(dados)
        
        # Ordenar por percentual
        resultado.sort(key=lambda x: x['percentual'], reverse=True)
        
        return resultado[:10]  # Top 10


class ExcelAvancadoExporter:
    """
    Classe para geração de arquivos Excel com formatação avançada.
    """
    
    def __init__(self):
        self.wb = None
        self.estilos = {}
        self._configurar_estilos()
    
    def _configurar_estilos(self):
        """Configura estilos para formatação profissional."""
        
        # Estilo de cabeçalho principal
        self.estilos['header_principal'] = NamedStyle(name="header_principal")
        self.estilos['header_principal'].font = Font(
            bold=True, color="FFFFFF", size=12
        )
        self.estilos['header_principal'].fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        self.estilos['header_principal'].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.estilos['header_principal'].border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Estilo de cabeçalho secundário
        self.estilos['header_secundario'] = NamedStyle(name="header_secundario")
        self.estilos['header_secundario'].font = Font(
            bold=True, color="000000", size=10
        )
        self.estilos['header_secundario'].fill = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )
        self.estilos['header_secundario'].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.estilos['header_secundario'].border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Estilo de dados
        self.estilos['dados'] = NamedStyle(name="dados")
        self.estilos['dados'].font = Font(size=9)
        self.estilos['dados'].alignment = Alignment(
            horizontal="left", vertical="center"
        )
        self.estilos['dados'].border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Estilo numérico
        self.estilos['numerico'] = NamedStyle(name="numerico")
        self.estilos['numerico'].font = Font(size=9)
        self.estilos['numerico'].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.estilos['numerico'].border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Estilo percentual
        self.estilos['percentual'] = NamedStyle(name="percentual")
        self.estilos['percentual'].font = Font(size=9, bold=True)
        self.estilos['percentual'].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.estilos['percentual'].number_format = '0.00%'
        self.estilos['percentual'].border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
    
    def gerar_excel_consolidado_geral(self, dados: Dict[str, Any], config: Dict[str, Any]) -> HttpResponse:
        """Gera Excel para consolidado geral."""
        
        self.wb = openpyxl.Workbook()
        
        # Remover planilha padrão
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Adicionar estilos ao workbook
        for estilo in self.estilos.values():
            if estilo.name not in [s.name for s in self.wb.named_styles]:
                self.wb.add_named_style(estilo)
        
        # Criar aba principal
        ws_principal = self.wb.create_sheet("Consolidado Geral")
        self._criar_aba_consolidado_principal(ws_principal, dados, config)
        
        if config.get('multiplas_abas'):
            # Aba de estatísticas
            ws_stats = self.wb.create_sheet("Estatísticas")
            self._criar_aba_estatisticas(ws_stats, dados['estatisticas'])
            
            # Aba de dados detalhados
            ws_detalhes = self.wb.create_sheet("Dados Detalhados")
            self._criar_aba_dados_detalhados(ws_detalhes, dados['presencas_detalhadas'])
        
        if config.get('incluir_graficos'):
            # Aba de gráficos
            ws_graficos = self.wb.create_sheet("Gráficos")
            self._criar_aba_graficos(ws_graficos, dados)
        
        # Gerar resposta HTTP
        return self._gerar_resposta_excel("consolidado_geral.xlsx")
    
    def _criar_aba_consolidado_principal(self, ws, dados: Dict[str, Any], config: Dict[str, Any]):
        """Cria aba principal do consolidado."""
        
        # Título
        titulo = config.get('titulo_personalizado', 'Relatório Consolidado de Presenças')
        ws.merge_cells('A1:P1')
        ws['A1'] = titulo
        ws['A1'].style = 'header_principal'
        ws.row_dimensions[1].height = 25
        
        # Informações do relatório
        row = 3
        ws[f'A{row}'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f'A{row}'].font = Font(size=9, italic=True)
        
        row += 1
        filtros = dados['filtros_aplicados']
        if filtros:
            ws[f'A{row}'] = "Filtros aplicados:"
            ws[f'A{row}'].font = Font(size=9, bold=True)
            row += 1
            
            for chave, valor in filtros.items():
                if valor:
                    ws[f'B{row}'] = f"{chave}: {valor}"
                    ws[f'B{row}'].font = Font(size=9)
                    row += 1
        
        # Estatísticas resumo
        row += 2
        stats = dados['estatisticas']
        
        # Criar tabela de estatísticas
        self._criar_tabela_estatisticas_resumo(ws, row, stats)
        
        # Dados consolidados
        row += 8
        self._criar_tabela_dados_consolidados(ws, row, dados)
    
    def _criar_tabela_estatisticas_resumo(self, ws, start_row: int, stats: Dict[str, Any]):
        """Cria tabela de estatísticas resumo."""
        
        # Cabeçalho
        ws[f'A{start_row}'] = "ESTATÍSTICAS GERAIS"
        ws[f'A{start_row}'].style = 'header_principal'
        ws.merge_cells(f'A{start_row}:D{start_row}')
        
        # Dados
        estatisticas_dados = [
            ["Total de Alunos", stats['total_alunos']],
            ["Total de Atividades", stats['total_atividades']],
            ["Total de Presenças", stats['total_presencas']],
            ["Total de Faltas", stats['total_faltas']],
            ["Percentual Médio", f"{stats['percentual_medio']:.2f}%"],
            ["Total de Voluntários", stats.get('voluntarios_total', 0)],
        ]
        
        for i, (label, valor) in enumerate(estatisticas_dados, start=start_row + 1):
            ws[f'A{i}'] = label
            ws[f'A{i}'].style = 'header_secundario'
            ws[f'B{i}'] = valor
            ws[f'B{i}'].style = 'numerico'
        
        # Formatação condicional para percentual
        percentual_cell = ws[f'B{start_row + 5}']
        if stats['percentual_medio'] >= 80:
            percentual_cell.fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid"
            )
        elif stats['percentual_medio'] >= 60:
            percentual_cell.fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )
        else:
            percentual_cell.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
    
    def _criar_tabela_dados_consolidados(self, ws, start_row: int, dados: Dict[str, Any]):
        """Cria tabela principal de dados consolidados."""
        
        presencas_detalhadas = dados['presencas_detalhadas']
        atividades = dados['atividades']
        
        # Cabeçalhos
        headers = ["Aluno", "Turma"]
        
        # Headers dinâmicos por atividade
        for atividade in atividades:
            headers.extend([
                f"{atividade.nome} - Conv",
                f"{atividade.nome} - Pres",
                f"{atividade.nome} - Faltas",
                f"{atividade.nome} - %",
            ])
        
        headers.extend(["Total Conv.", "Total Pres.", "Total Faltas", "% Médio"])
        
        # Escrever cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.style = 'header_principal'
        
        # Agrupar dados por aluno
        dados_por_aluno = self._agrupar_dados_por_aluno(presencas_detalhadas, atividades)
        
        # Escrever dados
        row = start_row + 1
        for aluno_dados in dados_por_aluno.values():
            col = 1
            
            # Dados básicos
            ws.cell(row=row, column=col, value=aluno_dados['aluno'].nome).style = 'dados'
            col += 1
            ws.cell(row=row, column=col, value=aluno_dados['turma'].nome).style = 'dados'
            col += 1
            
            # Dados por atividade
            for atividade in atividades:
                presenca = aluno_dados['atividades'].get(atividade.id)
                if presenca:
                    ws.cell(row=row, column=col, value=presenca.convocacoes).style = 'numerico'
                    ws.cell(row=row, column=col + 1, value=presenca.presencas).style = 'numerico'
                    ws.cell(row=row, column=col + 2, value=presenca.faltas).style = 'numerico'
                    
                    # Percentual com formatação condicional
                    perc_cell = ws.cell(
                        row=row, column=col + 3, 
                        value=float(presenca.percentual_presenca) / 100
                    )
                    perc_cell.style = 'percentual'
                    self._aplicar_formatacao_condicional_percentual(perc_cell)
                else:
                    for i in range(4):
                        ws.cell(row=row, column=col + i, value=0).style = 'numerico'
                col += 4
            
            # Totais
            totais = aluno_dados['totais']
            ws.cell(row=row, column=col, value=totais['convocacoes']).style = 'numerico'
            ws.cell(row=row, column=col + 1, value=totais['presencas']).style = 'numerico'
            ws.cell(row=row, column=col + 2, value=totais['faltas']).style = 'numerico'
            
            # Percentual médio
            perc_medio = (
                totais['presencas'] / totais['convocacoes'] 
                if totais['convocacoes'] > 0 else 0
            )
            perc_cell = ws.cell(row=row, column=col + 3, value=perc_medio)
            perc_cell.style = 'percentual'
            self._aplicar_formatacao_condicional_percentual(perc_cell)
            
            row += 1
        
        # Ajustar largura das colunas
        self._ajustar_largura_colunas(ws)
        
        # Criar tabela formatada
        self._criar_tabela_formatada(ws, start_row, row - 1, len(headers))
    
    def _agrupar_dados_por_aluno(self, presencas_detalhadas, atividades) -> Dict[int, Dict]:
        """Agrupa dados de presença por aluno."""
        
        dados_por_aluno = {}
        
        for presenca in presencas_detalhadas:
            aluno_id = presenca.aluno.id
            if aluno_id not in dados_por_aluno:
                dados_por_aluno[aluno_id] = {
                    'aluno': presenca.aluno,
                    'turma': presenca.turma,
                    'atividades': {},
                    'totais': {
                        'convocacoes': 0,
                        'presencas': 0,
                        'faltas': 0,
                    },
                }
            
            dados_por_aluno[aluno_id]['atividades'][presenca.atividade.id] = presenca
            dados_por_aluno[aluno_id]['totais']['convocacoes'] += presenca.convocacoes
            dados_por_aluno[aluno_id]['totais']['presencas'] += presenca.presencas
            dados_por_aluno[aluno_id]['totais']['faltas'] += presenca.faltas
        
        return dados_por_aluno
    
    def _aplicar_formatacao_condicional_percentual(self, cell):
        """Aplica formatação condicional baseada no percentual."""
        
        valor = cell.value if cell.value else 0
        
        if valor >= 0.8:  # >= 80%
            cell.fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid"
            )
        elif valor >= 0.6:  # >= 60%
            cell.fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )
        else:  # < 60%
            cell.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
    
    def _ajustar_largura_colunas(self, ws):
        """Ajusta largura das colunas automaticamente."""
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _criar_tabela_formatada(self, ws, start_row: int, end_row: int, num_cols: int):
        """Cria tabela formatada com estilo do Excel."""
        
        # Definir referência da tabela
        ref = f"A{start_row}:{get_column_letter(num_cols)}{end_row}"
        
        # Criar tabela
        table = Table(displayName="TabelaConsolidado", ref=ref)
        
        # Estilo da tabela
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Adicionar tabela à planilha
        ws.add_table(table)
    
    def _criar_aba_estatisticas(self, ws, estatisticas: Dict[str, Any]):
        """Cria aba de estatísticas detalhadas."""
        
        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = "ESTATÍSTICAS DETALHADAS"
        ws['A1'].style = 'header_principal'
        
        row = 3
        
        # Resumo geral
        self._criar_secao_estatisticas(ws, row, "Resumo Geral", [
            ["Total de Alunos", estatisticas['total_alunos']],
            ["Total de Atividades", estatisticas['total_atividades']],
            ["Total de Presenças", estatisticas['total_presencas']],
            ["Total de Faltas", estatisticas['total_faltas']],
            ["Percentual Médio", f"{estatisticas['percentual_medio']:.2f}%"],
        ])
        
        self._ajustar_largura_colunas(ws)
    
    def _criar_secao_estatisticas(self, ws, start_row: int, titulo: str, dados: List[List]):
        """Cria uma seção de estatísticas."""
        
        # Título da seção
        ws[f'A{start_row}'] = titulo
        ws[f'A{start_row}'].style = 'header_secundario'
        ws.merge_cells(f'A{start_row}:B{start_row}')
        
        # Dados
        for i, (label, valor) in enumerate(dados, start=start_row + 1):
            ws[f'A{i}'] = label
            ws[f'A{i}'].style = 'dados'
            ws[f'B{i}'] = valor
            ws[f'B{i}'].style = 'numerico'
        
        return start_row + len(dados) + 2
    
    def _criar_aba_dados_detalhados(self, ws, presencas_detalhadas):
        """Cria aba com dados detalhados linha por linha."""
        
        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = "DADOS DETALHADOS POR PRESENÇA"
        ws['A1'].style = 'header_principal'
        
        # Cabeçalhos
        headers = [
            "Aluno", "Turma", "Atividade", "Data", 
            "Convocações", "Presenças", "Faltas", "Percentual"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = 'header_secundario'
        
        # Dados
        row = 4
        for presenca in presencas_detalhadas:
            ws.cell(row=row, column=1, value=presenca.aluno.nome).style = 'dados'
            ws.cell(row=row, column=2, value=presenca.turma.nome).style = 'dados'
            ws.cell(row=row, column=3, value=presenca.atividade.nome).style = 'dados'
            ws.cell(row=row, column=4, value=presenca.atividade.data).style = 'dados'
            ws.cell(row=row, column=5, value=presenca.convocacoes).style = 'numerico'
            ws.cell(row=row, column=6, value=presenca.presencas).style = 'numerico'
            ws.cell(row=row, column=7, value=presenca.faltas).style = 'numerico'
            
            perc_cell = ws.cell(
                row=row, column=8, 
                value=float(presenca.percentual_presenca) / 100
            )
            perc_cell.style = 'percentual'
            self._aplicar_formatacao_condicional_percentual(perc_cell)
            
            row += 1
        
        # Criar tabela
        if row > 4:
            self._criar_tabela_formatada(ws, 3, row - 1, len(headers))
        
        self._ajustar_largura_colunas(ws)
    
    def _criar_aba_graficos(self, ws, dados: Dict[str, Any]):
        """Cria aba com gráficos."""
        
        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = "GRÁFICOS E ANÁLISES"
        ws['A1'].style = 'header_principal'
        
        # Preparar dados para gráfico de barras - estatísticas gerais
        stats = dados['estatisticas']
        
        # Dados para gráfico
        ws['A3'] = "Estatística"
        ws['B3'] = "Valor"
        ws['A3'].style = 'header_secundario'
        ws['B3'].style = 'header_secundario'
        
        grafico_dados = [
            ["Presenças", stats['total_presencas']],
            ["Faltas", stats['total_faltas']],
            ["Voluntários", stats.get('voluntarios_total', 0)],
        ]
        
        for i, (label, valor) in enumerate(grafico_dados, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = valor
            ws[f'A{i}'].style = 'dados'
            ws[f'B{i}'].style = 'numerico'
        
        # Criar gráfico de barras
        chart = BarChart()
        chart.title = "Distribuição Geral de Presenças"
        chart.y_axis.title = "Quantidade"
        chart.x_axis.title = "Tipo"
        
        # Definir dados do gráfico
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(grafico_dados))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(grafico_dados))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Posicionar gráfico
        ws.add_chart(chart, "D3")
        
        # Gráfico de pizza para percentuais
        row_pizza = 3 + len(grafico_dados) + 5
        
        ws[f'A{row_pizza}'] = "Distribuição Percentual"
        ws[f'A{row_pizza}'].style = 'header_secundario'
        ws.merge_cells(f'A{row_pizza}:B{row_pizza}')
        
        row_pizza += 1
        ws[f'A{row_pizza}'] = "Presenças"
        ws[f'B{row_pizza}'] = stats['total_presencas']
        ws[f'A{row_pizza + 1}'] = "Faltas"
        ws[f'B{row_pizza + 1}'] = stats['total_faltas']
        
        # Criar gráfico de pizza
        pie_chart = PieChart()
        pie_chart.title = "Distribuição Presença vs Faltas"
        
        pie_data = Reference(ws, min_col=2, min_row=row_pizza, max_row=row_pizza + 1)
        pie_cats = Reference(ws, min_col=1, min_row=row_pizza, max_row=row_pizza + 1)
        
        pie_chart.add_data(pie_data)
        pie_chart.set_categories(pie_cats)
        
        # Adicionar rótulos de dados
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        
        # Posicionar gráfico de pizza
        ws.add_chart(pie_chart, f"D{row_pizza}")
    
    def _gerar_resposta_excel(self, nome_arquivo: str) -> HttpResponse:
        """Gera resposta HTTP com arquivo Excel."""
        
        # Salvar em buffer
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        # Criar resposta
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        return response


# Continua na próxima parte do arquivo...
