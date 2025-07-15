"""
Views simplificadas para exportação avançada de dados de presenças.
Versão que funciona sem dependências externas opcionais.
"""

import csv
import io
import logging
from datetime import datetime, timedelta
from typing import Dict, Any

from django.http import HttpResponse, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.generic import TemplateView
from django.utils import timezone
from django.db.models import Sum, Count

from ..models import PresencaDetalhada
from turmas.models import Turma
from .consolidado import ConsolidadoPresencasView

logger = logging.getLogger(__name__)


class ExportacaoAvancadaView(LoginRequiredMixin, TemplateView):
    """
    View principal para exportação avançada de dados.
    Interface para seleção de formato e configurações.
    """
    template_name = 'presencas/exportacao/exportacao_avancada.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Formatos disponíveis (versão simplificada)
        context['formatos'] = [
            {'valor': 'csv', 'nome': 'CSV (.csv)'},
            {'valor': 'excel_basico', 'nome': 'Excel Básico (.xlsx)'},
        ]
        
        # Templates de relatório
        context['templates'] = [
            {'valor': 'consolidado_geral', 'nome': 'Consolidado Geral'},
            {'valor': 'por_turma', 'nome': 'Relatório por Turma'},
            {'valor': 'estatisticas_executivas', 'nome': 'Estatísticas Executivas'},
        ]
        
        # Opções de período
        context['periodos'] = [
            {'valor': 'atual', 'nome': 'Período Atual'},
            {'valor': 'ultimo_mes', 'nome': 'Último Mês'},
            {'valor': 'ultimo_trimestre', 'nome': 'Último Trimestre'},
            {'valor': 'ano_atual', 'nome': 'Ano Atual'},
            {'valor': 'personalizado', 'nome': 'Período Personalizado'},
        ]
        
        # Filtros disponíveis
        context['turmas'] = Turma.objects.filter(ativa=True).order_by('nome')
        context['cursos'] = []  # Simplificado
        
        return context


class ProcessarExportacaoView(LoginRequiredMixin, View):
    """
    View para processar solicitações de exportação.
    """
    
    def post(self, request, *args, **kwargs):
        """Processa requisição de exportação."""
        try:
            formato = request.POST.get('formato', 'csv')
            template = request.POST.get('template', 'consolidado_geral')
            configuracoes = self._extrair_configuracoes(request)
            
            # Obter dados
            dados = self._obter_dados(template, configuracoes)
            
            # Gerar arquivo conforme formato
            if formato == 'csv':
                return self._gerar_csv(template, dados, configuracoes)
            elif formato == 'excel_basico':
                return self._gerar_excel_basico(template, dados, configuracoes)
            else:
                return JsonResponse({
                    'erro': f'Formato não suportado: {formato}'
                }, status=400)
            
        except Exception as e:
            logger.error(f"Erro ao processar exportação: {str(e)}")
            return JsonResponse({
                'erro': f'Erro interno: {str(e)}'
            }, status=500)
    
    def _extrair_configuracoes(self, request) -> Dict[str, Any]:
        """Extrai configurações da requisição."""
        return {
            'periodo': request.POST.get('periodo', 'atual'),
            'data_inicio': request.POST.get('data_inicio'),
            'data_fim': request.POST.get('data_fim'),
            'turma_id': request.POST.get('turma_id'),
            'titulo_personalizado': request.POST.get('titulo_personalizado', ''),
        }
    
    def _obter_dados(self, template: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados conforme template e configurações."""
        
        # Criar instância do consolidado para reaproveitar lógica
        consolidado_view = ConsolidadoPresencasView()
        consolidado_view.request = self.request
        
        # Aplicar filtros baseados nas configurações
        filtros = self._construir_filtros(config)
        
        if template == 'consolidado_geral':
            return self._obter_dados_consolidado_geral(filtros)
        elif template == 'por_turma':
            return self._obter_dados_por_turma(filtros)
        elif template == 'estatisticas_executivas':
            return self._obter_estatisticas_executivas(filtros)
        else:
            return {'erro': f'Template não suportado: {template}'}
    
    def _construir_filtros(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Constrói filtros a partir das configurações."""
        filtros = {}
        
        # Filtros de período
        if config.get('periodo') != 'personalizado':
            if config['periodo'] == 'ultimo_mes':
                filtros['data_inicio'] = timezone.now() - timedelta(days=30)
            elif config['periodo'] == 'ultimo_trimestre':
                filtros['data_inicio'] = timezone.now() - timedelta(days=90)
            elif config['periodo'] == 'ano_atual':
                filtros['data_inicio'] = timezone.now().replace(month=1, day=1)
        else:
            if config.get('data_inicio'):
                filtros['data_inicio'] = datetime.strptime(
                    config['data_inicio'], '%Y-%m-%d'
                ).date()
            if config.get('data_fim'):
                filtros['data_fim'] = datetime.strptime(
                    config['data_fim'], '%Y-%m-%d'
                ).date()
        
        # Filtros de turma
        if config.get('turma_id'):
            filtros['turma_id'] = config['turma_id']
        
        return filtros
    
    def _obter_dados_consolidado_geral(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados para relatório consolidado geral."""
        
        # Reutilizar lógica do consolidado existente
        consolidado_view = ConsolidadoPresencasView()
        consolidado_view.request = self.request
        
        try:
            presencas_detalhadas = consolidado_view.obter_presencas_consolidadas(filtros)
            atividades = consolidado_view.obter_atividades_filtradas(filtros)
        except Exception as e:
            logger.error(f"Erro ao obter dados consolidados: {str(e)}")
            # Fallback para dados básicos
            query = PresencaDetalhada.objects.select_related('aluno', 'turma', 'atividade')
            if filtros.get('data_inicio'):
                query = query.filter(atividade__data__gte=filtros['data_inicio'])
            if filtros.get('data_fim'):
                query = query.filter(atividade__data__lte=filtros['data_fim'])
            if filtros.get('turma_id'):
                query = query.filter(turma_id=filtros['turma_id'])
            
            presencas_detalhadas = list(query[:1000])  # Limitar para performance
            atividades = []
        
        # Calcular estatísticas gerais
        estatisticas = self._calcular_estatisticas_gerais(presencas_detalhadas)
        
        return {
            'presencas_detalhadas': presencas_detalhadas,
            'atividades': atividades,
            'estatisticas': estatisticas,
            'filtros_aplicados': filtros,
        }
    
    def _obter_dados_por_turma(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém dados agrupados por turma."""
        
        # Base query
        query = PresencaDetalhada.objects.select_related('aluno', 'turma', 'atividade')
        
        # Aplicar filtros
        if filtros.get('data_inicio'):
            query = query.filter(atividade__data__gte=filtros['data_inicio'])
        if filtros.get('data_fim'):
            query = query.filter(atividade__data__lte=filtros['data_fim'])
        if filtros.get('turma_id'):
            query = query.filter(turma_id=filtros['turma_id'])
        
        # Agrupar por turma
        dados_por_turma = {}
        for presenca in query[:1000]:  # Limitar para performance
            turma_id = presenca.turma.id
            if turma_id not in dados_por_turma:
                dados_por_turma[turma_id] = {
                    'turma': presenca.turma,
                    'presencas': [],
                    'estatisticas': {
                        'total_alunos': set(),
                        'total_presencas': 0,
                        'total_faltas': 0,
                        'percentual_medio': 0,
                    }
                }
            
            dados_por_turma[turma_id]['presencas'].append(presenca)
            dados_por_turma[turma_id]['estatisticas']['total_alunos'].add(presenca.aluno.id)
            dados_por_turma[turma_id]['estatisticas']['total_presencas'] += presenca.presencas
            dados_por_turma[turma_id]['estatisticas']['total_faltas'] += presenca.faltas
        
        # Finalizar estatísticas
        for turma_data in dados_por_turma.values():
            stats = turma_data['estatisticas']
            stats['total_alunos'] = len(stats['total_alunos'])
            total_convocacoes = stats['total_presencas'] + stats['total_faltas']
            if total_convocacoes > 0:
                stats['percentual_medio'] = (
                    stats['total_presencas'] / total_convocacoes * 100
                )
        
        return {
            'dados_por_turma': dados_por_turma,
            'filtros_aplicados': filtros,
        }
    
    def _obter_estatisticas_executivas(self, filtros: Dict[str, Any]) -> Dict[str, Any]:
        """Obtém estatísticas executivas simplificadas."""
        
        # Dados base
        query = PresencaDetalhada.objects.select_related('aluno', 'turma', 'atividade')
        
        # Aplicar filtros
        if filtros.get('data_inicio'):
            query = query.filter(atividade__data__gte=filtros['data_inicio'])
        if filtros.get('data_fim'):
            query = query.filter(atividade__data__lte=filtros['data_fim'])
        
        # Estatísticas agregadas
        agregados = query.aggregate(
            total_presencas=Sum('presencas'),
            total_faltas=Sum('faltas'),
            total_alunos=Count('aluno', distinct=True),
            total_atividades=Count('atividade', distinct=True),
        )
        
        total_conv = (agregados['total_presencas'] or 0) + (agregados['total_faltas'] or 0)
        percentual_geral = (
            (agregados['total_presencas'] or 0) / total_conv * 100 
            if total_conv > 0 else 0
        )
        
        estatisticas = {
            'resumo_geral': {
                'total_alunos': agregados['total_alunos'] or 0,
                'total_atividades': agregados['total_atividades'] or 0,
                'total_presencas': agregados['total_presencas'] or 0,
                'total_faltas': agregados['total_faltas'] or 0,
                'percentual_geral': round(percentual_geral, 2),
            }
        }
        
        return {
            'estatisticas_executivas': estatisticas,
            'filtros_aplicados': filtros,
        }
    
    def _calcular_estatisticas_gerais(self, presencas_detalhadas) -> Dict[str, Any]:
        """Calcula estatísticas gerais dos dados."""
        
        if not presencas_detalhadas:
            return {
                'total_alunos': 0,
                'total_atividades': 0,
                'total_presencas': 0,
                'total_faltas': 0,
                'percentual_medio': 0,
            }
        
        alunos_unicos = set()
        atividades_unicas = set()
        total_presencas = 0
        total_faltas = 0
        
        for presenca in presencas_detalhadas:
            alunos_unicos.add(presenca.aluno.id)
            atividades_unicas.add(presenca.atividade.id)
            total_presencas += presenca.presencas
            total_faltas += presenca.faltas
        
        total_convocacoes = total_presencas + total_faltas
        percentual_medio = (
            (total_presencas / total_convocacoes * 100) if total_convocacoes > 0 else 0
        )
        
        return {
            'total_alunos': len(alunos_unicos),
            'total_atividades': len(atividades_unicas),
            'total_presencas': total_presencas,
            'total_faltas': total_faltas,
            'percentual_medio': round(percentual_medio, 2),
        }
    
    def _gerar_csv(self, template: str, dados: Dict[str, Any], config: Dict[str, Any]) -> HttpResponse:
        """Gera arquivo CSV do relatório."""
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="relatorio_presencas.csv"'
        
        # Adicionar BOM para Excel
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')  # Usar ; para Excel brasileiro
        
        # Cabeçalho com informações
        titulo = config.get('titulo_personalizado', 'Relatório de Presenças')
        writer.writerow([titulo])
        writer.writerow([f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        writer.writerow([])
        
        if template == 'consolidado_geral':
            self._escrever_csv_consolidado_geral(writer, dados)
        elif template == 'por_turma':
            self._escrever_csv_por_turma(writer, dados)
        elif template == 'estatisticas_executivas':
            self._escrever_csv_estatisticas(writer, dados)
        
        return response
    
    def _escrever_csv_consolidado_geral(self, writer, dados: Dict[str, Any]):
        """Escreve dados do consolidado geral no CSV."""
        
        # Estatísticas resumo
        stats = dados['estatisticas']
        writer.writerow(['ESTATÍSTICAS GERAIS'])
        writer.writerow(['Total de Alunos', stats['total_alunos']])
        writer.writerow(['Total de Atividades', stats['total_atividades']])
        writer.writerow(['Total de Presenças', stats['total_presencas']])
        writer.writerow(['Total de Faltas', stats['total_faltas']])
        writer.writerow(['Percentual Médio', f"{stats['percentual_medio']:.2f}%"])
        writer.writerow([])
        
        # Cabeçalhos da tabela de dados
        writer.writerow([
            'Aluno', 'Turma', 'Atividade', 'Data', 
            'Convocações', 'Presenças', 'Faltas', 'Percentual'
        ])
        
        # Dados detalhados
        for presenca in dados['presencas_detalhadas']:
            writer.writerow([
                presenca.aluno.nome,
                presenca.turma.nome,
                presenca.atividade.nome,
                presenca.atividade.data.strftime('%d/%m/%Y'),
                presenca.convocacoes,
                presenca.presencas,
                presenca.faltas,
                f"{float(presenca.percentual_presenca):.2f}%"
            ])
    
    def _escrever_csv_por_turma(self, writer, dados: Dict[str, Any]):
        """Escreve dados agrupados por turma no CSV."""
        
        for turma_data in dados['dados_por_turma'].values():
            turma = turma_data['turma']
            stats = turma_data['estatisticas']
            
            writer.writerow([f"TURMA: {turma.nome}"])
            writer.writerow(['Total de Alunos', stats['total_alunos']])
            writer.writerow(['Total de Presenças', stats['total_presencas']])
            writer.writerow(['Total de Faltas', stats['total_faltas']])
            writer.writerow(['Percentual Médio', f"{stats['percentual_medio']:.2f}%"])
            writer.writerow([])
            
            # Dados detalhados da turma
            writer.writerow(['Aluno', 'Atividade', 'Data', 'Presenças', 'Faltas', 'Percentual'])
            
            for presenca in turma_data['presencas']:
                writer.writerow([
                    presenca.aluno.nome,
                    presenca.atividade.nome,
                    presenca.atividade.data.strftime('%d/%m/%Y'),
                    presenca.presencas,
                    presenca.faltas,
                    f"{float(presenca.percentual_presenca):.2f}%"
                ])
            
            writer.writerow([])
            writer.writerow(['-' * 50])
            writer.writerow([])
    
    def _escrever_csv_estatisticas(self, writer, dados: Dict[str, Any]):
        """Escreve estatísticas executivas no CSV."""
        
        stats = dados['estatisticas_executivas']['resumo_geral']
        
        writer.writerow(['ESTATÍSTICAS EXECUTIVAS'])
        writer.writerow([])
        writer.writerow(['Total de Alunos', stats['total_alunos']])
        writer.writerow(['Total de Atividades', stats['total_atividades']])
        writer.writerow(['Total de Presenças', stats['total_presencas']])
        writer.writerow(['Total de Faltas', stats['total_faltas']])
        writer.writerow(['Percentual Geral', f"{stats['percentual_geral']:.2f}%"])
    
    def _gerar_excel_basico(self, template: str, dados: Dict[str, Any], config: Dict[str, Any]) -> HttpResponse:
        """Gera arquivo Excel básico usando openpyxl se disponível."""
        
        # Verificar se openpyxl está disponível
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            # Fallback para CSV se Excel não estiver disponível
            return self._gerar_csv(template, dados, config)
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Presenças"
        
        # Título
        titulo = config.get('titulo_personalizado', 'Relatório de Presenças')
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Data de geração
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True)
        
        row = 4
        
        if template == 'consolidado_geral':
            row = self._escrever_excel_consolidado_geral(ws, dados, row)
        elif template == 'por_turma':
            row = self._escrever_excel_por_turma(ws, dados, row)
        elif template == 'estatisticas_executivas':
            row = self._escrever_excel_estatisticas(ws, dados, row)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Criar resposta
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_presencas.xlsx"'
        
        return response
    
    def _escrever_excel_consolidado_geral(self, ws, dados: Dict[str, Any], start_row: int) -> int:
        """Escreve dados do consolidado geral no Excel."""
        
        # Estatísticas
        stats = dados['estatisticas']
        ws[f'A{start_row}'] = "ESTATÍSTICAS GERAIS"
        # Aplicar estilo apenas se disponível
        try:
            from openpyxl.styles import Font
            ws[f'A{start_row}'].font = Font(bold=True)
        except ImportError:
            pass
        
        start_row += 1
        stats_data = [
            ['Total de Alunos', stats['total_alunos']],
            ['Total de Atividades', stats['total_atividades']],
            ['Total de Presenças', stats['total_presencas']],
            ['Total de Faltas', stats['total_faltas']],
            ['Percentual Médio', f"{stats['percentual_medio']:.2f}%"],
        ]
        
        for label, valor in stats_data:
            ws[f'A{start_row}'] = label
            ws[f'B{start_row}'] = valor
            start_row += 1
        
        start_row += 2
        
        # Cabeçalhos da tabela
        headers = ['Aluno', 'Turma', 'Atividade', 'Data', 'Convocações', 'Presenças', 'Faltas', 'Percentual']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            # Aplicar estilos apenas se disponível
            try:
                from openpyxl.styles import Font, PatternFill
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            except ImportError:
                pass
        
        start_row += 1
        
        # Dados
        for presenca in dados['presencas_detalhadas']:
            ws.cell(row=start_row, column=1, value=presenca.aluno.nome)
            ws.cell(row=start_row, column=2, value=presenca.turma.nome)
            ws.cell(row=start_row, column=3, value=presenca.atividade.nome)
            ws.cell(row=start_row, column=4, value=presenca.atividade.data.strftime('%d/%m/%Y'))
            ws.cell(row=start_row, column=5, value=presenca.convocacoes)
            ws.cell(row=start_row, column=6, value=presenca.presencas)
            ws.cell(row=start_row, column=7, value=presenca.faltas)
            ws.cell(row=start_row, column=8, value=f"{float(presenca.percentual_presenca):.2f}%")
            start_row += 1
        
        return start_row
    
    def _escrever_excel_por_turma(self, ws, dados: Dict[str, Any], start_row: int) -> int:
        """Escreve dados agrupados por turma no Excel."""
        
        for turma_data in dados['dados_por_turma'].values():
            turma = turma_data['turma']
            stats = turma_data['estatisticas']
            
            # Título da turma
            ws[f'A{start_row}'] = f"TURMA: {turma.nome}"
            # Aplicar estilo apenas se disponível
            try:
                from openpyxl.styles import Font
                ws[f'A{start_row}'].font = Font(bold=True, size=12)
            except ImportError:
                pass
            start_row += 1
            
            # Estatísticas da turma
            turma_stats = [
                ['Total de Alunos', stats['total_alunos']],
                ['Total de Presenças', stats['total_presencas']],
                ['Total de Faltas', stats['total_faltas']],
                ['Percentual Médio', f"{stats['percentual_medio']:.2f}%"],
            ]
            
            for label, valor in turma_stats:
                ws[f'A{start_row}'] = label
                ws[f'B{start_row}'] = valor
                start_row += 1
            
            start_row += 1
            
            # Cabeçalhos
            headers = ['Aluno', 'Atividade', 'Data', 'Presenças', 'Faltas', 'Percentual']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                # Aplicar estilo apenas se disponível
                try:
                    from openpyxl.styles import Font
                    cell.font = Font(bold=True)
                except ImportError:
                    pass
            
            start_row += 1
            
            # Dados da turma
            for presenca in turma_data['presencas']:
                ws.cell(row=start_row, column=1, value=presenca.aluno.nome)
                ws.cell(row=start_row, column=2, value=presenca.atividade.nome)
                ws.cell(row=start_row, column=3, value=presenca.atividade.data.strftime('%d/%m/%Y'))
                ws.cell(row=start_row, column=4, value=presenca.presencas)
                ws.cell(row=start_row, column=5, value=presenca.faltas)
                ws.cell(row=start_row, column=6, value=f"{float(presenca.percentual_presenca):.2f}%")
                start_row += 1
            
            start_row += 2
        
        return start_row
    
    def _escrever_excel_estatisticas(self, ws, dados: Dict[str, Any], start_row: int) -> int:
        """Escreve estatísticas executivas no Excel."""
        
        stats = dados['estatisticas_executivas']['resumo_geral']
        
        ws[f'A{start_row}'] = "ESTATÍSTICAS EXECUTIVAS"
        # Aplicar estilo apenas se disponível
        try:
            from openpyxl.styles import Font
            ws[f'A{start_row}'].font = Font(bold=True, size=12)
        except ImportError:
            pass
        start_row += 2
        
        stats_data = [
            ['Total de Alunos', stats['total_alunos']],
            ['Total de Atividades', stats['total_atividades']],
            ['Total de Presenças', stats['total_presencas']],
            ['Total de Faltas', stats['total_faltas']],
            ['Percentual Geral', f"{stats['percentual_geral']:.2f}%"],
        ]
        
        for label, valor in stats_data:
            ws[f'A{start_row}'] = label
            ws[f'B{start_row}'] = valor
            start_row += 1
        
        return start_row


class GerenciarAgendamentosView(LoginRequiredMixin, TemplateView):
    """
    View para gerenciar agendamentos de relatórios.
    Versão simplificada sem agendamento automático.
    """
    template_name = 'presencas/exportacao/gerenciar_agendamentos.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['agendamentos'] = []  # Simplificado
        context['mensagem'] = "Funcionalidade de agendamento será implementada em versão futura."
        
        return context


def agendamento_form_ajax(request):
    """View simplificada para formulário de agendamento."""
    return JsonResponse({
        'html': '<p class="text-muted">Funcionalidade em desenvolvimento.</p>'
    })
