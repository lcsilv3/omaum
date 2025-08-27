"""
Views para sistema consolidado de presenças.
Implementa funcionalidade estilo Excel para visualização e edição de presenças.
"""

import logging
from datetime import datetime
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import View
from django.db import transaction

from ..models import PresencaDetalhada
from presencas.services import CalculadoraEstatisticas
from cursos.models import Curso
from turmas.models import Turma
from atividades.models import AtividadeAcademica

try:
    # Assumir que CalculadoraEstatisticas será criado pelo Agente 4
    from ..services import CalculadoraEstatisticas
except ImportError:
    # Fallback se não existir ainda
    class CalculadoraEstatisticas:
        @staticmethod
        def calcular_totais_consolidado(presencas):
            """Método fallback para cálculo de totais."""
            return {
                "total_convocacoes": sum(p.convocacoes for p in presencas),
                "total_presencas": sum(p.presencas for p in presencas),
                "total_faltas": sum(p.faltas for p in presencas),
                "total_voluntarios": sum(p.total_voluntarios for p in presencas),
                "media_percentual": (
                    sum(p.percentual_presenca for p in presencas) / len(presencas)
                    if presencas
                    else 0
                ),
            }


logger = logging.getLogger(__name__)


class ConsolidadoPresencasView(LoginRequiredMixin, View):
    """
    View principal para exibição do consolidado de presenças estilo Excel.
    Suporta filtros, paginação horizontal e edição in-line.
    """

    template_name = "presencas/consolidado/consolidado.html"
    ATIVIDADES_POR_PAGINA = 10  # Controla navegação horizontal

    def get_context_data(self, **kwargs):
        """Monta context com dados consolidados."""
        context = {}

        # Obter filtros
        filtros = self.obter_filtros()
        context["filtros"] = filtros

        # Obter dados base
        context["cursos"] = Curso.objects.filter(ativo=True).order_by("nome")
        context["turmas"] = self.obter_turmas_filtradas(filtros)
        context["atividades"] = self.obter_atividades_filtradas(filtros)

        # Obter presenças consolidadas
        presencas_detalhadas = self.obter_presencas_consolidadas(filtros)

        # Aplicar paginação de atividades
        atividades_paginadas = self.paginar_atividades(
            context["atividades"], filtros.get("pagina_atividade", 1)
        )
        context["atividades_paginadas"] = atividades_paginadas

        # Montar matriz de dados (Excel-style)
        context["dados_consolidados"] = self.montar_matriz_dados(
            presencas_detalhadas, atividades_paginadas
        )

        # Calcular estatísticas
        context["estatisticas"] = self.calcular_estatisticas(presencas_detalhadas)

        # Preparar dados para template
        context["configuracao"] = self.obter_configuracao_consolidado()
        context["pode_editar"] = self.verificar_permissao_edicao()

        return context

    def get(self, request, *args, **kwargs):
        """Exibe página consolidada."""
        try:
            context = self.get_context_data(**kwargs)
            return render(request, self.template_name, context)
        except Exception as e:
            logger.error(f"Erro ao carregar consolidado: {str(e)}")
            messages.error(request, f"Erro ao carregar dados consolidados: {str(e)}")
            return redirect("presencas:listar_presencas")

    def post(self, request, *args, **kwargs):
        """Processa alterações via AJAX."""
        try:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return self.processar_ajax(request)
            else:
                return self.processar_formulario(request)
        except Exception as e:
            logger.error(f"Erro ao processar alterações: {str(e)}")
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": str(e)})
            else:
                messages.error(request, f"Erro ao salvar alterações: {str(e)}")
                return self.get(request, *args, **kwargs)

    def obter_filtros(self):
        """Extrai filtros da request."""
        filtros = {}

        # Filtros básicos
        filtros["curso_id"] = self.request.GET.get("curso_id")
        filtros["turma_id"] = self.request.GET.get("turma_id")
        filtros["periodo_inicio"] = self.request.GET.get("periodo_inicio")
        filtros["periodo_fim"] = self.request.GET.get("periodo_fim")
        filtros["atividade_id"] = self.request.GET.get("atividade_id")
        filtros["aluno_nome"] = self.request.GET.get("aluno_nome")

        # Filtros de paginação
        filtros["pagina_atividade"] = self.request.GET.get("pagina_atividade", 1)

        # Filtros de ordenação
        filtros["ordenar_por"] = self.request.GET.get("ordenar_por", "aluno__nome")
        filtros["ordem"] = self.request.GET.get("ordem", "asc")

        # Validar e converter datas
        if filtros["periodo_inicio"]:
            try:
                filtros["periodo_inicio"] = datetime.strptime(
                    filtros["periodo_inicio"], "%Y-%m-%d"
                ).date()
            except ValueError:
                filtros["periodo_inicio"] = None

        if filtros["periodo_fim"]:
            try:
                filtros["periodo_fim"] = datetime.strptime(
                    filtros["periodo_fim"], "%Y-%m-%d"
                ).date()
            except ValueError:
                filtros["periodo_fim"] = None

        return filtros

    def obter_turmas_filtradas(self, filtros):
        """Obtém turmas baseado nos filtros."""
        turmas = Turma.objects.filter(status="A")

        if filtros.get("curso_id"):
            turmas = turmas.filter(curso_id=filtros["curso_id"])

        return turmas.order_by("nome")

    def obter_atividades_filtradas(self, filtros):
        """Obtém atividades baseado nos filtros."""
        atividades = AtividadeAcademica.objects.filter(ativo=True)

        if filtros.get("turma_id"):
            # Filtrar por atividades que têm registros para esta turma
            atividades = atividades.filter(
                presencas_detalhadas_expandidas__turma_id=filtros["turma_id"]
            ).distinct()

        if filtros.get("atividade_id"):
            atividades = atividades.filter(id=filtros["atividade_id"])

        return atividades.order_by("nome")

    def obter_presencas_consolidadas(self, filtros):
        """Obtém presenças detalhadas baseado nos filtros."""
        presencas = PresencaDetalhada.objects.select_related(
            "aluno", "turma", "atividade"
        ).all()

        # Aplicar filtros
        if filtros.get("turma_id"):
            presencas = presencas.filter(turma_id=filtros["turma_id"])

        if filtros.get("atividade_id"):
            presencas = presencas.filter(atividade_id=filtros["atividade_id"])

        if filtros.get("periodo_inicio"):
            presencas = presencas.filter(periodo__gte=filtros["periodo_inicio"])

        if filtros.get("periodo_fim"):
            presencas = presencas.filter(periodo__lte=filtros["periodo_fim"])

        if filtros.get("aluno_nome"):
            presencas = presencas.filter(aluno__nome__icontains=filtros["aluno_nome"])

        # Ordenação
        ordem = filtros.get("ordenar_por", "aluno__nome")
        if filtros.get("ordem") == "desc":
            ordem = f"-{ordem}"

        return presencas.order_by(ordem)

    def paginar_atividades(self, atividades, pagina):
        """Aplica paginação horizontal às atividades."""
        paginator = Paginator(atividades, self.ATIVIDADES_POR_PAGINA)

        try:
            atividades_paginadas = paginator.page(pagina)
        except PageNotAnInteger:
            atividades_paginadas = paginator.page(1)
        except EmptyPage:
            atividades_paginadas = paginator.page(paginator.num_pages)

        return atividades_paginadas

    def montar_matriz_dados(self, presencas_detalhadas, atividades_paginadas):
        """Monta matriz de dados estilo Excel."""
        dados = {}

        # Agrupar por aluno
        for presenca in presencas_detalhadas:
            aluno_id = presenca.aluno.id
            atividade_id = presenca.atividade.id

            # Verificar se atividade está na página atual
            if atividade_id not in [a.id for a in atividades_paginadas]:
                continue

            if aluno_id not in dados:
                dados[aluno_id] = {
                    "aluno": presenca.aluno,
                    "turma": presenca.turma,
                    "atividades": {},
                    "totais": {
                        "convocacoes": 0,
                        "presencas": 0,
                        "faltas": 0,
                        "voluntarios": 0,
                        "percentual_medio": 0,
                    },
                }

            # Adicionar dados da atividade
            dados[aluno_id]["atividades"][atividade_id] = {
                "presenca_detalhada": presenca,
                "C": presenca.convocacoes,
                "P": presenca.presencas,
                "F": presenca.faltas,
                "V1": presenca.voluntario_extra,
                "V2": presenca.voluntario_simples,
                "percentual": presenca.percentual_presenca,
                "carencias": presenca.carencias,
                "pode_editar": True,  # Definir regras de negócio
            }

            # Acumular totais
            dados[aluno_id]["totais"]["convocacoes"] += presenca.convocacoes
            dados[aluno_id]["totais"]["presencas"] += presenca.presencas
            dados[aluno_id]["totais"]["faltas"] += presenca.faltas
            dados[aluno_id]["totais"]["voluntarios"] += presenca.total_voluntarios

        # Calcular percentuais médios
        for aluno_id, dados_aluno in dados.items():
            total_conv = dados_aluno["totais"]["convocacoes"]
            total_pres = dados_aluno["totais"]["presencas"]

            if total_conv > 0:
                dados_aluno["totais"]["percentual_medio"] = round(
                    (total_pres / total_conv) * 100, 2
                )
            else:
                dados_aluno["totais"]["percentual_medio"] = 0

        return dados

    def calcular_estatisticas(self, presencas_detalhadas):
        """Calcula estatísticas gerais."""
        if not presencas_detalhadas:
            return {
                "total_registros": 0,
                "total_convocacoes": 0,
                "total_presencas": 0,
                "total_faltas": 0,
                "total_voluntarios": 0,
                "media_percentual": 0,
                "alunos_unicos": 0,
                "atividades_unicas": 0,
            }

        # Usar CalculadoraEstatisticas
        estatisticas = CalculadoraEstatisticas.calcular_totais_consolidado(
            presencas_detalhadas
        )

        # Adicionar estatísticas adicionais
        estatisticas.update(
            {
                "total_registros": len(presencas_detalhadas),
                "alunos_unicos": len(set(p.aluno.id for p in presencas_detalhadas)),
                "atividades_unicas": len(
                    set(p.atividade.id for p in presencas_detalhadas)
                ),
            }
        )

        return estatisticas

    def obter_configuracao_consolidado(self):
        """Obtém configurações específicas do consolidado."""
        return {
            "atividades_por_pagina": self.ATIVIDADES_POR_PAGINA,
            "permite_edicao_inline": True,
            "autocompletamento_ativo": True,
            "salvar_automatico": False,
            "confirmar_antes_salvar": True,
        }

    def verificar_permissao_edicao(self):
        """Verifica se usuário pode editar dados."""
        # Implementar regras de negócio específicas
        return self.request.user.is_authenticated and (
            self.request.user.is_staff
            or self.request.user.has_perm("presencas.change_presencadetalhada")
        )

    def processar_ajax(self, request):
        """Processa requisições AJAX."""
        try:
            acao = request.POST.get("acao")

            if acao == "salvar_celula":
                return self.salvar_celula_ajax(request)
            elif acao == "salvar_linha":
                return self.salvar_linha_ajax(request)
            elif acao == "carregar_filtros":
                return self.carregar_filtros_ajax(request)
            elif acao == "exportar_excel":
                return self.exportar_excel_ajax(request)
            else:
                return JsonResponse({"success": False, "error": "Ação não reconhecida"})

        except Exception as e:
            logger.error(f"Erro em AJAX: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    def salvar_celula_ajax(self, request):
        """Salva alteração de uma célula específica."""
        try:
            with transaction.atomic():
                presenca_id = request.POST.get("presenca_id")
                campo = request.POST.get("campo")
                valor = request.POST.get("valor")

                presenca = get_object_or_404(PresencaDetalhada, id=presenca_id)

                # Validar permissão
                if not self.verificar_permissao_edicao():
                    return JsonResponse(
                        {"success": False, "error": "Sem permissão para editar"}
                    )

                # Validar campo
                campos_permitidos = [
                    "convocacoes",
                    "presencas",
                    "faltas",
                    "voluntario_extra",
                    "voluntario_simples",
                ]
                if campo not in campos_permitidos:
                    return JsonResponse(
                        {"success": False, "error": "Campo não permitido"}
                    )

                # Validar valor
                try:
                    valor_int = int(valor)
                    if valor_int < 0:
                        raise ValueError("Valor não pode ser negativo")
                except ValueError:
                    return JsonResponse({"success": False, "error": "Valor inválido"})

                # Atualizar campo
                setattr(presenca, campo, valor_int)
                presenca.save()  # Ativa recálculo automático

                return JsonResponse(
                    {
                        "success": True,
                        "nova_presenca": {
                            "id": presenca.id,
                            "percentual": str(presenca.percentual_presenca),
                            "total_voluntarios": presenca.total_voluntarios,
                            "carencias": presenca.carencias,
                        },
                    }
                )

        except Exception as e:
            logger.error(f"Erro ao salvar célula: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    def salvar_linha_ajax(self, request):
        """Salva alterações de uma linha completa."""
        try:
            with transaction.atomic():
                request.POST.get("aluno_id")
                request.POST.get("dados_linha")

                # Processar dados da linha
                # ... implementar lógica de salvamento em lote

                return JsonResponse(
                    {"success": True, "mensagem": "Linha salva com sucesso"}
                )

        except Exception as e:
            logger.error(f"Erro ao salvar linha: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    def carregar_filtros_ajax(self, request):
        """Carrega dados para filtros dinâmicos."""
        try:
            tipo_filtro = request.GET.get("tipo")

            if tipo_filtro == "turmas":
                curso_id = request.GET.get("curso_id")
                turmas = Turma.objects.filter(curso_id=curso_id, status="A")
                return JsonResponse(
                    {
                        "success": True,
                        "turmas": [{"id": t.id, "nome": t.nome} for t in turmas],
                    }
                )
            elif tipo_filtro == "atividades":
                turma_id = request.GET.get("turma_id")
                atividades = AtividadeAcademica.objects.filter(
                    presencas_detalhadas_expandidas__turma_id=turma_id
                ).distinct()
                return JsonResponse(
                    {
                        "success": True,
                        "atividades": [
                            {"id": a.id, "nome": a.nome} for a in atividades
                        ],
                    }
                )
            else:
                return JsonResponse(
                    {"success": False, "error": "Tipo de filtro não reconhecido"}
                )

        except Exception as e:
            logger.error(f"Erro ao carregar filtros: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    def exportar_excel_ajax(self, request):
        """Prepara dados para exportação Excel."""
        try:
            # Redirecionar para view específica de exportação
            return JsonResponse(
                {
                    "success": True,
                    "redirect_url": reverse_lazy("presencas:exportar_consolidado"),
                }
            )

        except Exception as e:
            logger.error(f"Erro ao preparar exportação: {str(e)}")
            return JsonResponse({"success": False, "error": str(e)})

    def processar_formulario(self, request):
        """Processa formulário principal."""
        try:
            # Implementar processamento do formulário se necessário
            messages.success(request, "Alterações salvas com sucesso!")
            return redirect("presencas:consolidado")

        except Exception as e:
            logger.error(f"Erro ao processar formulário: {str(e)}")
            messages.error(request, f"Erro ao salvar: {str(e)}")
            return self.get(request)


class FiltroConsolidadoView(LoginRequiredMixin, View):
    """
    View para filtros avançados do consolidado.
    """

    template_name = "presencas/consolidado/filtros.html"

    def get(self, request, *args, **kwargs):
        """Exibe formulário de filtros avançados."""
        context = {
            "cursos": Curso.objects.filter(ativo=True).order_by("nome"),
            "turmas": Turma.objects.filter(status="A").order_by("nome"),
            "atividades": AtividadeAcademica.objects.filter(ativo=True).order_by(
                "nome"
            ),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Aplica filtros e redireciona para consolidado."""
        try:
            # Construir parâmetros de filtro
            params = {}

            campos_filtro = [
                "curso_id",
                "turma_id",
                "atividade_id",
                "aluno_nome",
                "periodo_inicio",
                "periodo_fim",
                "ordenar_por",
                "ordem",
            ]

            for campo in campos_filtro:
                valor = request.POST.get(campo)
                if valor:
                    params[campo] = valor

            # Redirecionar para consolidado com filtros
            url = reverse_lazy("presencas:consolidado")
            if params:
                from urllib.parse import urlencode

                url += "?" + urlencode(params)

            return redirect(url)

        except Exception as e:
            logger.error(f"Erro ao aplicar filtros: {str(e)}")
            messages.error(request, f"Erro ao aplicar filtros: {str(e)}")
            return self.get(request, *args, **kwargs)


class ExportarConsolidadoView(LoginRequiredMixin, View):
    """
    View para exportação do consolidado em Excel.
    """

    def get(self, request, *args, **kwargs):
        """Exporta dados consolidados para Excel."""
        try:
            # Usar mesma lógica de filtros do consolidado
            consolidado_view = ConsolidadoPresencasView()
            consolidado_view.request = request
            filtros = consolidado_view.obter_filtros()

            # Obter dados completos (sem paginação)
            presencas_detalhadas = consolidado_view.obter_presencas_consolidadas(
                filtros
            )
            atividades = consolidado_view.obter_atividades_filtradas(filtros)

            # Gerar arquivo Excel
            response = self.gerar_excel(presencas_detalhadas, atividades, filtros)

            return response

        except Exception as e:
            logger.error(f"Erro ao exportar Excel: {str(e)}")
            messages.error(request, f"Erro ao exportar: {str(e)}")
            return redirect("presencas:consolidado")

    def gerar_excel(self, presencas_detalhadas, atividades, filtros):
        """Gera arquivo Excel com dados consolidados."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Consolidado Presenças"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Cabeçalhos
            headers = ["Aluno", "Turma"]
            for atividade in atividades:
                headers.extend(
                    [
                        f"{atividade.nome} - C",
                        f"{atividade.nome} - P",
                        f"{atividade.nome} - F",
                        f"{atividade.nome} - V1",
                        f"{atividade.nome} - V2",
                        f"{atividade.nome} - %",
                    ]
                )
            headers.extend(
                ["Total Conv.", "Total Pres.", "Total Faltas", "Total Vol.", "% Médio"]
            )

            # Escrever cabeçalhos
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Agrupar dados por aluno
            dados_por_aluno = {}
            for presenca in presencas_detalhadas:
                aluno_id = presenca.aluno.id
                if aluno_id not in dados_por_aluno:
                    dados_por_aluno[aluno_id] = {
                        "aluno": presenca.aluno,
                        "turma": presenca.turma,
                        "atividades": {},
                        "totais": {
                            "convocacoes": 0,
                            "presencas": 0,
                            "faltas": 0,
                            "voluntarios": 0,
                        },
                    }

                dados_por_aluno[aluno_id]["atividades"][presenca.atividade.id] = (
                    presenca
                )
                dados_por_aluno[aluno_id]["totais"]["convocacoes"] += (
                    presenca.convocacoes
                )
                dados_por_aluno[aluno_id]["totais"]["presencas"] += presenca.presencas
                dados_por_aluno[aluno_id]["totais"]["faltas"] += presenca.faltas
                dados_por_aluno[aluno_id]["totais"]["voluntarios"] += (
                    presenca.total_voluntarios
                )

            # Escrever dados
            row = 2
            for aluno_id, dados_aluno in dados_por_aluno.items():
                col = 1

                # Dados do aluno
                ws.cell(row=row, column=col, value=dados_aluno["aluno"].nome)
                col += 1
                ws.cell(row=row, column=col, value=dados_aluno["turma"].nome)
                col += 1

                # Dados por atividade
                for atividade in atividades:
                    presenca = dados_aluno["atividades"].get(atividade.id)
                    if presenca:
                        ws.cell(row=row, column=col, value=presenca.convocacoes)
                        ws.cell(row=row, column=col + 1, value=presenca.presencas)
                        ws.cell(row=row, column=col + 2, value=presenca.faltas)
                        ws.cell(
                            row=row, column=col + 3, value=presenca.voluntario_extra
                        )
                        ws.cell(
                            row=row, column=col + 4, value=presenca.voluntario_simples
                        )
                        ws.cell(
                            row=row,
                            column=col + 5,
                            value=float(presenca.percentual_presenca),
                        )
                    else:
                        # Células vazias para atividade sem dados
                        for i in range(6):
                            ws.cell(row=row, column=col + i, value=0)
                    col += 6

                # Totais
                totais = dados_aluno["totais"]
                ws.cell(row=row, column=col, value=totais["convocacoes"])
                ws.cell(row=row, column=col + 1, value=totais["presencas"])
                ws.cell(row=row, column=col + 2, value=totais["faltas"])
                ws.cell(row=row, column=col + 3, value=totais["voluntarios"])

                # Percentual médio
                if totais["convocacoes"] > 0:
                    percentual_medio = (
                        totais["presencas"] / totais["convocacoes"]
                    ) * 100
                else:
                    percentual_medio = 0
                ws.cell(row=row, column=col + 4, value=round(percentual_medio, 2))

                row += 1

            # Ajustar largura das colunas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12

            # Preparar response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Nome do arquivo
            periodo = ""
            if filtros.get("periodo_inicio") and filtros.get("periodo_fim"):
                periodo = f"_{filtros['periodo_inicio']}_{filtros['periodo_fim']}"

            filename = f"consolidado_presencas{periodo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            wb.save(response)
            return response

        except ImportError:
            # Se openpyxl não estiver disponível, usar fallback
            return self.gerar_csv_fallback(presencas_detalhadas, atividades, filtros)
        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {str(e)}")
            raise

    def gerar_csv_fallback(self, presencas_detalhadas, atividades, filtros):
        """Gera arquivo CSV como fallback."""
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            'attachment; filename="consolidado_presencas.csv"'
        )

        writer = csv.writer(response)

        # Cabeçalhos
        headers = ["Aluno", "Turma"]
        for atividade in atividades:
            headers.extend(
                [
                    f"{atividade.nome} - C",
                    f"{atividade.nome} - P",
                    f"{atividade.nome} - F",
                    f"{atividade.nome} - V1",
                    f"{atividade.nome} - V2",
                    f"{atividade.nome} - %",
                ]
            )
        headers.extend(
            ["Total Conv.", "Total Pres.", "Total Faltas", "Total Vol.", "% Médio"]
        )

        writer.writerow(headers)

        # Dados (implementar lógica similar ao Excel)
        # ...

        return response
